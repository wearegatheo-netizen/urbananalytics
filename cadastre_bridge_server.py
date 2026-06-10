import json
import math
import base64
import csv
import io
import os
import re
import mimetypes
import subprocess
import sys
import threading
import time
import urllib.parse
import urllib.request
import zipfile
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from xml.etree import ElementTree as ET

from PIL import Image, ImageDraw, ImageFont


ROOT = Path(__file__).resolve().parent
APP_DIR = ROOT / "dist" / "app"
CADASTRE_SCRIPT = APP_DIR / "run_cadastre_browser_full.ps1"
DEFAULT_HTML = ROOT / "map_bisan_1056_6_satellite.html"
DEFAULT_OUT_DIR = Path.home() / "OneDrive" / "Desktop" / "연속지적"
VWORLD_GEOCODE_URL = "https://api.vworld.kr/req/address"
NOMINATIM_URL = "https://nominatim.openstreetmap.org/search"
OVERPASS_URL = "https://overpass-api.de/api/interpreter"
ESRI_TILE_URL = "https://server.arcgisonline.com/ArcGIS/rest/services/World_Imagery/MapServer/tile/{z}/{y}/{x}"
OSM_TILE_URL = "https://tile.openstreetmap.org/{z}/{x}/{y}.png"
# OSM 공식 타일서버는 서버측 대량 요청을 차단(Access blocked 타일)하므로
# 일반 지도는 OSM 데이터 기반의 Carto Voyager 타일을 사용한다.
CARTO_VOYAGER_TILE_URL = "https://a.basemaps.cartocdn.com/rastertiles/voyager/{z}/{x}/{y}.png"
CARTO_LIGHT_NOLABEL_TILE_URL = "https://a.basemaps.cartocdn.com/light_nolabels/{z}/{x}/{y}.png"

JOBS = {}

# 클라우드(컨테이너) 배포 모드: PORT 환경변수가 있거나 비-Windows이면 웹 모드로 간주.
# 웹 모드에서는 폴더 선택 대화상자·QGIS 연속지적 실행 등 데스크톱 전용 기능을 비활성화한다.
WEB_MODE = bool(os.environ.get("PORT")) or os.name != "nt"

# 요청별 컨텍스트(스레드-로컬): VWorld WFS(2D데이터 API)는 호출 도메인을 검증하므로,
# 페이지에 접속한 도메인(요청 Host/Origin)을 WFS 요청의 DOMAIN 파라미터로 자동 전달한다.
_request_ctx = threading.local()


def current_vworld_domain():
    # VWORLD_DOMAIN 환경변수가 있으면 최우선(터널/프록시 환경에서 키 등록 도메인 고정용).
    return os.environ.get("VWORLD_DOMAIN") or getattr(_request_ctx, "domain", "") or ""


def cors_headers(handler, status=200, content_type="application/json"):
    handler.send_response(status)
    handler.send_header("Access-Control-Allow-Origin", "*")
    handler.send_header("Access-Control-Allow-Methods", "GET, POST, OPTIONS")
    handler.send_header("Access-Control-Allow-Headers", "Content-Type")
    if content_type.startswith("text/") or content_type == "application/json":
        content_type = f"{content_type}; charset=utf-8"
    handler.send_header("Content-Type", content_type)


def json_response(handler, payload, status=200):
    data = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    cors_headers(handler, status=status)
    handler.send_header("Content-Length", str(len(data)))
    handler.end_headers()
    handler.wfile.write(data)


def http_get_json(url, params, retries=4):
    # VWorld는 짧은 시간에 호출이 몰리면 일시 차단(INCORRECT_KEY=XML)·과부하 응답을 준다.
    # 비정상 응답이면 백오프 후 재시도하여 버스트 차단을 흡수한다.
    full = f"{url}?{urllib.parse.urlencode(params)}"
    last = None
    for attempt in range(retries):
        try:
            req = urllib.request.Request(full, headers={"User-Agent": "Mozilla/5.0"})
            with urllib.request.urlopen(req, timeout=30) as res:
                text = res.read().decode("utf-8", errors="replace")
            return json.loads(text)
        except Exception as exc:
            last = exc
            if attempt < retries - 1:
                time.sleep(0.5 * (attempt + 1))
    raise last


def geocode(address, api_key):
    normalized = address.replace(" ", "")
    if "비산동1056" in normalized or "관악대로254" in normalized:
        return 37.401514, 126.9513167, "경기 안양시 동안구 관악대로 254 일원"
    if "서울송파구중대로113" in normalized or "송파구중대로113" in normalized:
        return 37.4939, 127.1198, "서울특별시 송파구 중대로 113"

    if api_key:
        last_error = None
        for addr_type in ("ROAD", "PARCEL"):
            params = {
                "service": "address",
                "request": "getcoord",
                "version": "2.0",
                "crs": "epsg:4326",
                "refine": "true",
                "simple": "false",
                "format": "json",
                "type": addr_type,
                "address": address,
                "key": api_key,
            }
            data = http_get_json(VWORLD_GEOCODE_URL, params)
            response = data.get("response", {})
            if response.get("status") == "OK":
                result = response["result"]
                point = result["point"]
                refined = result.get("refined", {}).get("text") or address
                return float(point["y"]), float(point["x"]), refined
            last_error = response
        raise ValueError(f"VWorld 주소 변환 실패: {last_error}")

    params = {"q": address, "format": "jsonv2", "limit": "1", "countrycodes": "kr"}
    req = urllib.request.Request(
        f"{NOMINATIM_URL}?{urllib.parse.urlencode(params)}",
        headers={"User-Agent": "CadastreMapBridge/1.0"},
    )
    with urllib.request.urlopen(req, timeout=30) as res:
        data = json.loads(res.read().decode("utf-8", errors="replace"))
    if data:
        result = data[0]
        return float(result["lat"]), float(result["lon"]), result.get("display_name") or address
    raise ValueError("주소를 좌표로 변환하지 못했습니다. VWorld API 키를 입력하거나 도로명 주소로 다시 시도해 주세요.")


def reverse_geocode_parcel(lat, lon, api_key):
    if not api_key:
        raise ValueError("지도에서 지번을 선택하려면 VWorld API 키가 필요합니다.")
    params = {
        "service": "address",
        "request": "getAddress",
        "version": "2.0",
        "crs": "epsg:4326",
        "point": f"{lon},{lat}",
        "format": "json",
        "type": "PARCEL",
        "key": api_key,
    }
    data = http_get_json(VWORLD_GEOCODE_URL, params)
    response = data.get("response", {})
    if response.get("status") != "OK":
        raise ValueError(f"VWorld 지번 조회 실패: {response}")
    result = response.get("result", [])
    if isinstance(result, dict):
        result = [result]
    for item in result:
        text = item.get("text") or item.get("address", {}).get("parcel") or item.get("address", {}).get("text")
        if text:
            return text
    raise ValueError("선택한 위치의 지번을 찾지 못했습니다.")


def http_get_bytes(url, params, retries=3, timeout=30):
    full = f"{url}?{urllib.parse.urlencode(params)}"
    last = None
    for attempt in range(retries):
        try:
            req = urllib.request.Request(full, headers={"User-Agent": "Mozilla/5.0"})
            with urllib.request.urlopen(req, timeout=timeout) as res:
                return res.read(), res.headers.get("content-type", "")
        except Exception as exc:
            last = exc
            if attempt < retries - 1:
                time.sleep(0.4 * (attempt + 1))
    raise last


VWORLD_WMS_URL = "https://api.vworld.kr/req/wms"


def _zoning_name_at(common, px, py, fallback_lonlat=None):
    """lt_c_uq111 WMS GetFeatureInfo로 픽셀 위치의 용도지역 명칭을 얻는다."""
    params = dict(common)
    params.update({
        "REQUEST": "GetFeatureInfo", "LAYERS": "lt_c_uq111", "QUERY_LAYERS": "lt_c_uq111",
        "STYLES": "lt_c_uq111", "INFO_FORMAT": "application/json",
        "I": str(int(px)), "J": str(int(py)), "FEATURE_COUNT": "1",
        "FORMAT": "image/png", "TRANSPARENT": "true",
    })
    name = ""
    try:
        data = http_get_json(VWORLD_WMS_URL, params)
        feats = data.get("features") or []
        if feats:
            props = feats[0].get("properties") or {}
            for k in ("dgm_nm", "prpos_area_dstrc_code_nm", "uname", "ucode_nm", "label", "dgm_nm_1"):
                v = str(props.get(k) or "").strip()
                if v:
                    name = v
                    break
            if not name:
                for v in props.values():
                    s = str(v or "").strip()
                    if "지역" in s or "지구" in s or "구역" in s:
                        name = s
                        break
    except Exception:
        name = ""
    return name


def sample_zoning_legend(api_key, bbox, width=320, height=320, max_items=16):
    """현재 보기 영역의 용도지역 WMS를 렌더해 실제 색↔명칭 범례를 만든다(화면과 일치)."""
    if not api_key:
        return []
    from PIL import Image
    import numpy as np
    minx, miny, maxx, maxy = (float(v) for v in bbox)
    if not (maxx > minx and maxy > miny):
        return []
    dom = current_vworld_domain()
    common = {
        "SERVICE": "WMS", "VERSION": "1.3.0", "CRS": "EPSG:4326",
        "BBOX": f"{miny},{minx},{maxy},{maxx}",   # 1.3.0 + EPSG:4326 → lat,lon 순서
        "WIDTH": str(width), "HEIGHT": str(height),
        "key": api_key, "DOMAIN": dom,
    }
    map_params = dict(common)
    map_params.update({
        "REQUEST": "GetMap", "LAYERS": "lt_c_uq111", "STYLES": "lt_c_uq111",
        "FORMAT": "image/png", "TRANSPARENT": "false", "BGCOLOR": "0xFFFFFF",
    })
    raw, ct = http_get_bytes(VWORLD_WMS_URL, map_params)
    if "image" not in (ct or ""):
        return []
    arr = np.asarray(Image.open(io.BytesIO(raw)).convert("RGB"))
    h, w, _ = arr.shape
    flat = arr.reshape(-1, 3).astype(int)
    r, g, b = flat[:, 0], flat[:, 1], flat[:, 2]
    mx = np.maximum(np.maximum(r, g), b)
    mn = np.minimum(np.minimum(r, g), b)
    is_white = (r >= 243) & (g >= 243) & (b >= 243)
    is_black = mx <= 55
    is_gray = ((mx - mn) <= 12) & (mx < 235)        # 경계선·라벨(회색) 제외
    keep = ~(is_white | is_black | is_gray)
    keep_pos = np.nonzero(keep)[0]
    if keep_pos.size == 0:
        return []
    cols = flat[keep_pos]
    # 정확 실제색(양자화 없음) 빈도 — 용도지역 채움은 단색 평면이므로 최빈 정확색이 곧 채움색
    ucol, ucnt = np.unique(cols, axis=0, return_counts=True)
    order = np.argsort(-ucnt)
    ar = arr[:, :, 0]; ag = arr[:, :, 1]; ab = arr[:, :, 2]
    min_px = max(40, int(h * w * 0.0015))
    legend, seen = [], set()
    for oi in order:
        if len(legend) >= max_items:
            break
        if ucnt[oi] < min_px:                         # 단색 채움이 충분히 큰 색만(혼합/경계 잡색 제외)
            continue
        cr, cg, cb = (int(v) for v in ucol[oi])
        # 정확히 이 채움색인 픽셀의 2D 마스크
        mask = (ar == cr) & (ag == cg) & (ab == cb)
        # 반복 4-이웃 침식 → '가장 깊은 내부' 픽셀(영역 코어). 같은 색이 여러 곳에
        # 흩어져 있어도 코어에서 조회하므로 경계 이웃 구역 오인식이 없다.
        deepest = mask
        cur = mask
        while True:
            e = cur.copy()
            e[1:, :] &= cur[:-1, :]
            e[:-1, :] &= cur[1:, :]
            e[:, 1:] &= cur[:, :-1]
            e[:, :-1] &= cur[:, 1:]
            if not e.any():
                break
            deepest = e
            cur = e
        ys, xs = np.nonzero(deepest)
        if ys.size == 0:
            continue
        cy, cx = ys.mean(), xs.mean()
        j = int(np.argmin((ys - cy) ** 2 + (xs - cx) ** 2))
        py, px = int(ys[j]), int(xs[j])
        name = _zoning_name_at(common, px, py)
        if not name or name in seen:
            continue
        seen.add(name)
        hexc = "#%02x%02x%02x" % (cr, cg, cb)
        legend.append({"name": name, "color": hexc})
    # 면적(빈도) 순 유지
    return legend


def reverse_geocode_both(lat, lon, api_key):
    """역지오코딩으로 (지번주소, 도로명주소) 반환. 실패 항목은 빈 문자열."""
    if not api_key:
        return "", ""
    params = {
        "service": "address", "request": "getAddress", "version": "2.0",
        "crs": "epsg:4326", "point": f"{lon},{lat}", "format": "json",
        "type": "BOTH", "key": api_key,
    }
    try:
        data = http_get_json(VWORLD_GEOCODE_URL, params)
    except Exception:
        return "", ""
    response = data.get("response", {})
    if response.get("status") != "OK":
        return "", ""
    result = response.get("result", [])
    if isinstance(result, dict):
        result = [result]
    parcel = road = ""
    for item in result:
        t = str(item.get("type") or "").lower()
        text = (item.get("text") or "").strip()
        if not text:
            continue
        if t == "parcel" and not parcel:
            parcel = text
        elif t == "road" and not road:
            road = text
    return parcel, road


def point_in_ring(lon, lat, ring):
    inside = False
    count = len(ring)
    if count < 4:
        return False
    j = count - 1
    for i in range(count):
        xi, yi = float(ring[i][0]), float(ring[i][1])
        xj, yj = float(ring[j][0]), float(ring[j][1])
        if ((yi > lat) != (yj > lat)) and (lon < (xj - xi) * (lat - yi) / (yj - yi) + xi):
            inside = not inside
        j = i
    return inside


def feature_contains_point(feature, lon, lat):
    # 외곽링/홀을 even-odd 규칙으로 함께 판정(홀 안의 점은 제외)
    inside = False
    for ring in geometry_rings(feature.get("geometry") or {}):
        if point_in_ring(lon, lat, ring):
            inside = not inside
    return inside


def feature_max_ring_area(feature):
    areas = [
        abs(projected_polygon_area(project_ring(ring)))
        for ring in geometry_rings(feature.get("geometry") or {})
        if len(ring) >= 4
    ]
    return max(areas) if areas else 0.0


def feature_centroid_distance(feature, lon, lat):
    points = [pt for ring in geometry_rings(feature.get("geometry") or {}) for pt in ring]
    if not points:
        return float("inf")
    cx = sum(float(pt[0]) for pt in points) / len(points)
    cy = sum(float(pt[1]) for pt in points) / len(points)
    return (cx - lon) ** 2 + (cy - lat) ** 2


def deg_box(lat, lon, radius_m):
    dlat = radius_m / 111320.0
    dlon = radius_m / (111320.0 * max(math.cos(math.radians(lat)), 1e-9))
    return lon - dlon, lat - dlat, lon + dlon, lat + dlat


def build_cadastre_dxf(address, api_key, radius, domain=""):
    """반경 내 연속지적(필지)+용도지역을 VWorld WFS로 모아 EPSG:5186 좌표의 DXF로 생성(웹용, QGIS 불필요)."""
    if not api_key:
        raise ValueError("연속지적 DXF에는 VWorld API 키가 필요합니다.")
    import ezdxf
    from ezdxf.enums import TextEntityAlignment
    from pyproj import Transformer

    radius = max(50.0, min(parse_float(radius, 300.0) or 300.0, 1000.0))
    lat, lon, refined = geocode(address, api_key)
    minx, miny, maxx, maxy = deg_box(lat, lon, radius)
    tf = Transformer.from_crs("EPSG:4326", "EPSG:5186", always_xy=True)
    cx, cy = tf.transform(lon, lat)

    fetch_errors = []

    def fetch(typename, maxf):
        params = {
            "REQUEST": "GetFeature", "TYPENAME": typename, "VERSION": "1.1.0",
            "MAXFEATURES": str(maxf), "SRSNAME": "EPSG:4326", "OUTPUT": "json",
            "BBOX": f"{minx},{miny},{maxx},{maxy}", "KEY": api_key,
        }
        dom = domain or current_vworld_domain()
        if dom:
            params["DOMAIN"] = dom
        url = "https://api.vworld.kr/req/wfs?" + urllib.parse.urlencode(params)
        last_msg = ""
        for attempt in range(4):
            raw = ""
            try:
                req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
                with urllib.request.urlopen(req, timeout=30) as res:
                    raw = res.read().decode("utf-8", errors="replace")
                feats = json.loads(raw).get("features")
                return (feats or []), maxf
            except Exception as exc:
                match = re.search(r"<ServiceException[^>]*>(.*?)</ServiceException>", raw, re.S) if raw else None
                last_msg = match.group(1).strip() if match else f"{type(exc).__name__} {str(exc)[:140]}"
                if attempt < 3:
                    time.sleep(0.5 * (attempt + 1))
        fetch_errors.append(f"{typename}: {last_msg[:240]}")
        return [], maxf

    bubun, mb = fetch("lp_pa_cbnd_bubun", 1000)
    bonbun, _ = fetch("lp_pa_cbnd_bonbun", 1000)
    zones, _ = fetch("lt_c_uq111", 300)
    # 추가 레이어: 도시계획시설(도로/교통시설)
    plan_roads, _ = fetch("lt_c_upisuq151", 1000)        # 도시계획(도로)
    plan_transit, _ = fetch("lt_c_upisuq152", 800)       # 도시계획(교통시설)
    truncated = len(bubun) >= mb or len(bonbun) >= mb

    doc = ezdxf.new("R2010")
    try:
        doc.units = 6  # meters
    except Exception:
        pass
    for name, color in [("연속지적_필지", 7), ("지번", 3), ("용도지역", 5), ("용도지역명", 6),
                        ("도시계획_도로", 4), ("도시계획_교통시설", 30), ("도시계획시설명", 6),
                        ("기준점", 1)]:
        if name not in doc.layers:
            doc.layers.add(name, color=color)
    msp = doc.modelspace()

    def proj_ring(ring):
        return [tf.transform(float(p[0]), float(p[1])) for p in ring if len(p) >= 2]

    def add_feature(feature, line_layer, text_layer, label, text_h):
        rings = [r for r in geometry_rings(feature.get("geometry") or {}) if len(r) >= 4]
        if not rings:
            return None
        best, best_area = None, -1.0
        for ring in rings:
            pts = proj_ring(ring)
            if len(pts) < 3:
                continue
            msp.add_lwpolyline(pts, close=True, dxfattribs={"layer": line_layer})
            area = abs(projected_polygon_area(pts))
            if area > best_area:
                best_area, best = area, pts
        if best is None:
            return None
        ccx = sum(p[0] for p in best) / len(best)
        ccy = sum(p[1] for p in best) / len(best)
        if label:
            msp.add_text(label, height=text_h, dxfattribs={"layer": text_layer}).set_placement(
                (ccx, ccy), align=TextEntityAlignment.MIDDLE_CENTER
            )
        return (round(ccx, 1), round(ccy, 1))

    seen = set()
    parcel_count = 0
    for feature in bubun + bonbun:
        props = feature.get("properties") or {}
        jibun = str(props.get("jibun") or props.get("bonbun") or "").strip()
        key = add_feature(feature, "연속지적_필지", "지번", jibun, 1.5)
        if key is None or key in seen:
            continue
        seen.add(key)
        parcel_count += 1

    zone_count = 0
    for feature in zones:
        props = feature.get("properties") or {}
        uname = str(props.get("uname") or "").strip()
        if add_feature(feature, "용도지역", "용도지역명", uname, 3.0) is not None:
            zone_count += 1

    # 도시계획시설(도로·교통시설): dgm_nm(시설명) 라벨
    plan_count = 0
    for feature in plan_roads:
        props = feature.get("properties") or {}
        nm = str(props.get("dgm_nm") or "도로").strip()
        if add_feature(feature, "도시계획_도로", "도시계획시설명", nm, 2.5) is not None:
            plan_count += 1
    for feature in plan_transit:
        props = feature.get("properties") or {}
        nm = str(props.get("dgm_nm") or "교통시설").strip()
        if add_feature(feature, "도시계획_교통시설", "도시계획시설명", nm, 2.5) is not None:
            plan_count += 1

    if not parcel_count and not zone_count and fetch_errors:
        raise ValueError(
            "VWorld 데이터(WFS) 조회에 실패했습니다: " + fetch_errors[0]
            + " · 키의 '데이터 API(WFS)' 권한과 인증 도메인, 일일 사용량을 확인하세요."
        )

    msp.add_circle((cx, cy), radius, dxfattribs={"layer": "기준점"})
    msp.add_point((cx, cy), dxfattribs={"layer": "기준점"})
    msp.add_text(f"기준점: {refined} (반경 {int(radius)}m)", height=2.5,
                 dxfattribs={"layer": "기준점"}).set_placement((cx, cy + radius + 5))

    import tempfile
    path = tempfile.mktemp(suffix=".dxf")
    doc.saveas(path)
    with open(path, "rb") as fh:
        data = fh.read()
    try:
        os.remove(path)
    except OSError:
        pass
    return {"data": data, "parcelCount": parcel_count, "zoneCount": zone_count,
            "planCount": plan_count,
            "refined": refined, "radius": int(radius), "truncated": truncated,
            "errors": fetch_errors}


def get_parcel_feature(lat, lon, api_key):
    if not api_key:
        raise ValueError("지적도 도형 조회에는 VWorld API 키가 필요합니다.")
    last_error = None
    candidates = []
    # 본번/부번 두 레이어에서 클릭 지점 주변 필지들을 모은 뒤, 점을 '포함'하는 필지를 고른다.
    # (예전 방식은 MAXFEATURES=1 + 넓은 BBOX의 첫 피처를 써서 인접한 큰 도형(가구)을 잡았다.)
    for typename in ("lp_pa_cbnd_bubun", "lp_pa_cbnd_bonbun"):
        for delta in (0.00012, 0.0004):
            params = {
                "REQUEST": "GetFeature",
                "TYPENAME": typename,
                "VERSION": "1.1.0",
                "MAXFEATURES": "100",
                "SRSNAME": "EPSG:4326",
                "OUTPUT": "json",
                "BBOX": f"{lon - delta},{lat - delta},{lon + delta},{lat + delta}",
                "KEY": api_key,
                "DOMAIN": current_vworld_domain(),
            }
            try:
                data = http_get_json("https://api.vworld.kr/req/wfs", params)
            except Exception as exc:
                last_error = exc
                continue
            features = data.get("features") or []
            if not features:
                last_error = data
                continue
            containing = [f for f in features if feature_contains_point(f, lon, lat)]
            if containing:
                # 점을 포함하는 필지 중 가장 작은(=가장 구체적인) 필지를 선택
                best = min(containing, key=feature_max_ring_area)
                return {"type": "FeatureCollection", "features": [best]}
            candidates.extend(features)
            break  # 이 레이어에서 피처는 받았으므로(포함 필지는 없음) 다음 레이어로
    if candidates:
        # 포함 필지를 못 찾으면 클릭점에서 중심이 가장 가까운 필지로 대체
        best = min(candidates, key=lambda f: feature_centroid_distance(f, lon, lat))
        return {"type": "FeatureCollection", "features": [best]}
    raise ValueError(f"선택한 위치의 지적도 도형을 찾지 못했습니다: {last_error}")


def get_parcel_feature_by_pnu(pnu, api_key):
    """PNU(19자리)로 연속지적 필지 도형·속성을 조회한다. 못 찾으면 None."""
    pnu = re.sub(r"\D", "", str(pnu or ""))
    if len(pnu) != 19 or not api_key:
        return None
    # VWorld WFS는 attrFilter/CQL을 무시하므로 OGC FILTER(xml)로 pnu 등치 필터.
    # pnu는 숫자만 남겨 두었으므로 XML 이스케이프 불필요.
    flt = (f"<Filter><PropertyIsEqualTo><PropertyName>pnu</PropertyName>"
           f"<Literal>{pnu}</Literal></PropertyIsEqualTo></Filter>")
    last_error = None
    for typename in ("lp_pa_cbnd_bubun", "lp_pa_cbnd_bonbun"):
        params = {
            "REQUEST": "GetFeature",
            "TYPENAME": typename,
            "VERSION": "1.1.0",
            "MAXFEATURES": "5",
            "SRSNAME": "EPSG:4326",
            "OUTPUT": "json",
            "FILTER": flt,
            "KEY": api_key,
            "DOMAIN": current_vworld_domain(),
        }
        try:
            data = http_get_json("https://api.vworld.kr/req/wfs", params)
        except Exception as exc:
            last_error = exc
            continue
        features = data.get("features") or []
        if features:
            return {"type": "FeatureCollection", "features": [features[0]]}
        last_error = data
    return None


def geometry_rings(geometry):
    coords = (geometry or {}).get("coordinates") or []
    geom_type = (geometry or {}).get("type")
    if geom_type == "Polygon":
        return coords
    if geom_type == "MultiPolygon":
        return [ring for polygon in coords for ring in polygon]
    return []


def project_point(lon, lat, origin_lat):
    meters_per_lat = 111320.0
    meters_per_lon = 111320.0 * math.cos(math.radians(origin_lat))
    return lon * meters_per_lon, lat * meters_per_lat


def project_ring(ring):
    origin_lat = sum(float(pt[1]) for pt in ring) / len(ring)
    return [project_point(float(pt[0]), float(pt[1]), origin_lat) for pt in ring]


def projected_polygon_area(points):
    area = 0.0
    for idx, point in enumerate(points):
        nxt = points[(idx + 1) % len(points)]
        area += point[0] * nxt[1] - nxt[0] * point[1]
    return area / 2.0


def primary_ring(feature):
    rings = geometry_rings(feature.get("geometry") or {})
    rings = [ring for ring in rings if len(ring) >= 4]
    if not rings:
        raise ValueError("필지 도형 좌표를 읽지 못했습니다.")
    return max(rings, key=lambda ring: abs(projected_polygon_area(project_ring(ring))))


def feature_geom_area(feature):
    """필지 도형의 실제 면적(㎡, EPSG:5186): 모든 폴리곤의 (외곽링 − 구멍링) 합.
    멀티폴리곤(여러 조각)·구멍(도넛)을 정확히 반영한다(primary_ring 단일링 계산의 오류 보정)."""
    geom = feature.get("geometry") or {}
    gtype = geom.get("type")
    coords = geom.get("coordinates") or []
    if gtype == "Polygon":
        polygons = [coords]
    elif gtype == "MultiPolygon":
        polygons = coords
    else:
        return 0.0
    total = 0.0
    for poly in polygons:
        if not poly:
            continue
        outer = poly[0]
        outer_area = abs(projected_polygon_area(project_ring(outer))) if len(outer) >= 4 else 0.0
        holes = sum(abs(projected_polygon_area(project_ring(r))) for r in poly[1:] if len(r) >= 4)
        total += max(outer_area - holes, 0.0)
    return total


def lonlat_centroid(ring):
    return sum(float(pt[1]) for pt in ring) / len(ring), sum(float(pt[0]) for pt in ring) / len(ring)


def polygon_segments(points):
    return [(points[idx], points[(idx + 1) % len(points)]) for idx in range(len(points) - 1)]


def segment_length(a, b):
    return math.hypot(b[0] - a[0], b[1] - a[1])


def point_segment_distance(point, a, b):
    px, py = point
    ax, ay = a
    bx, by = b
    dx = bx - ax
    dy = by - ay
    denom = dx * dx + dy * dy
    if denom == 0:
        return math.hypot(px - ax, py - ay)
    t = max(0.0, min(1.0, ((px - ax) * dx + (py - ay) * dy) / denom))
    return math.hypot(px - (ax + t * dx), py - (ay + t * dy))


def segment_distance(a1, a2, b1, b2):
    return min(
        point_segment_distance(a1, b1, b2),
        point_segment_distance(a2, b1, b2),
        point_segment_distance(b1, a1, a2),
        point_segment_distance(b2, a1, a2),
    )


def parse_width_tag(value):
    if value is None:
        return None
    match = re.search(r"\d+(?:[\.,]\d+)?", str(value))
    return float(match.group(0).replace(",", ".")) if match else None


def estimate_road_width(tags):
    width = parse_width_tag(tags.get("width"))
    if width:
        return width
    lanes = parse_width_tag(tags.get("lanes"))
    if lanes:
        return max(4.0, lanes * 3.2)
    defaults = {
        "motorway": 28.0,
        "trunk": 24.0,
        "primary": 20.0,
        "secondary": 15.0,
        "tertiary": 12.0,
        "unclassified": 8.0,
        "residential": 8.0,
        "service": 6.0,
        "living_street": 6.0,
        "footway": 4.0,
    }
    return defaults.get(tags.get("highway"), 8.0)


def osm_roads_near(lat, lon, radius=120):
    query = f"""
    [out:json][timeout:15];
    way(around:{int(radius)},{lat},{lon})["highway"];
    (._;>;);
    out body;
    """
    req = urllib.request.Request(
        OVERPASS_URL,
        data=urllib.parse.urlencode({"data": query}).encode("utf-8"),
        headers={"User-Agent": "CadastreMapBridge/1.0"},
    )
    with urllib.request.urlopen(req, timeout=25) as res:
        data = json.loads(res.read().decode("utf-8", errors="replace"))
    nodes = {item["id"]: (float(item["lat"]), float(item["lon"])) for item in data.get("elements", []) if item.get("type") == "node"}
    roads = []
    for item in data.get("elements", []):
        tags = item.get("tags", {})
        if item.get("type") != "way" or "highway" not in tags:
            continue
        coords = [nodes[node_id] for node_id in item.get("nodes", []) if node_id in nodes]
        if len(coords) < 2:
            continue
        roads.append(
            {
                "name": tags.get("name") or tags.get("name:ko") or tags.get("highway") or "road",
                "highway": tags.get("highway"),
                "width": estimate_road_width(tags),
                "coords": coords,
            }
        )
    return roads


def analyze_roads_for_ring(ring, projected, area):
    lat, lon = lonlat_centroid(ring)
    parcel_segments = polygon_segments(projected)
    touching = []
    try:
        roads = osm_roads_near(lat, lon)
    except Exception:
        roads = []
    for road in roads:
        road_points = [project_point(coord[1], coord[0], lat) for coord in road["coords"]]
        min_dist = None
        frontage = 0.0
        for idx in range(len(road_points) - 1):
            r1, r2 = road_points[idx], road_points[idx + 1]
            for p1, p2 in parcel_segments:
                dist = segment_distance(p1, p2, r1, r2)
                min_dist = dist if min_dist is None else min(min_dist, dist)
                if dist <= 8.0:
                    frontage += segment_length(p1, p2)
        if min_dist is not None and min_dist <= 12.0:
            item = dict(road)
            item["distance"] = min_dist
            item["frontage"] = frontage
            touching.append(item)
    touching.sort(key=lambda road: (road["distance"], -road["frontage"]))
    if touching:
        main = touching[0]
        return {
            "roadName": main["name"],
            "roadWidth": main["width"],
            "frontage": max(main["frontage"], math.sqrt(area)),
            "touchingRoads": touching[:3],
        }
    return {
        "roadName": "접도 도로 자동판정 필요",
        "roadWidth": None,
        "frontage": math.sqrt(area),
        "touchingRoads": [],
    }


def coverage_source(payload):
    district = parse_float(payload.get("districtPlanCoverage"), None)
    ordinance = parse_float(payload.get("ordinanceCoverage"), None)
    default = parse_float(payload.get("coverageRatio"), 60.0)
    if district:
        return district, "지구단위계획 확인값"
    if ordinance:
        return ordinance, "조례 확인값"
    return default, "기본값"


# 국토계획법 시행령 제84조 용도지역별 건폐율 상한(%) — 조례 조회 실패 시 폴백
COVERAGE_CAP = [
    ("제1종전용주거", 50), ("제2종전용주거", 50),
    ("제1종일반주거", 60), ("제2종일반주거", 60), ("제3종일반주거", 50),
    ("준주거", 70),
    ("중심상업", 90), ("일반상업", 80), ("근린상업", 70), ("유통상업", 80),
    ("전용공업", 70), ("일반공업", 70), ("준공업", 70),
    ("보전녹지", 20), ("생산녹지", 20), ("자연녹지", 20),
    ("보전관리", 20), ("생산관리", 20), ("계획관리", 40),
    ("농림", 20), ("자연환경보전", 20),
]


def coverage_cap_for(zone_name):
    name = str(zone_name or "").replace(" ", "")
    for key, pct in COVERAGE_CAP:
        if key in name:
            return float(pct)
    return None


def get_landuse(lat, lon, api_key):
    """VWorld 용도지역(lt_c_uq111)에서 점이 포함된 용도지역과 지자체를 반환."""
    if not api_key:
        return None
    delta = 0.00015
    params = {
        "REQUEST": "GetFeature", "TYPENAME": "lt_c_uq111", "VERSION": "1.1.0",
        "MAXFEATURES": "30", "SRSNAME": "EPSG:4326", "OUTPUT": "json",
        "BBOX": f"{lon - delta},{lat - delta},{lon + delta},{lat + delta}", "KEY": api_key,
        "DOMAIN": current_vworld_domain(),
    }
    try:
        data = http_get_json("https://api.vworld.kr/req/wfs", params)
    except Exception:
        return None
    feats = data.get("features") or []
    if not feats:
        return None
    chosen = next((f for f in feats if feature_contains_point(f, lon, lat)), feats[0])
    props = chosen.get("properties") or {}
    return {
        "zone": props.get("uname"),
        "sido": props.get("sido_name"),
        "sigungu": props.get("sigg_name"),
        "sggcd": props.get("std_sggcd"),
    }


_ORDINANCE_CACHE = {}


def fetch_ordinance_coverage(sigungu, zone_name, oc):
    """법제처 자치법규 API로 해당 지자체 도시계획조례의 용도지역별 건폐율(%)을 조회."""
    if not oc or not sigungu or not zone_name:
        return None
    cache_key = (sigungu, zone_name)
    if cache_key in _ORDINANCE_CACHE:
        return _ORDINANCE_CACHE[cache_key]
    result = None
    try:
        # 1) 도시계획조례 검색 → 자치법규 일련번호(MST)
        search = http_get_json("https://www.law.go.kr/DRF/lawSearch.do", {
            "OC": oc, "target": "ordin", "type": "JSON",
            "query": f"{sigungu} 도시계획 조례", "display": "20",
        })
        block = search.get("OrdinSearch") or search.get("LawSearch") or {}
        laws = block.get("law") or block.get("ordin") or []
        if isinstance(laws, dict):
            laws = [laws]
        mst = None
        for law in laws:
            name = str(law.get("자치법규명") or law.get("법령명한글") or law.get("법령명") or "")
            if "도시계획" in name and "조례" in name and "시행" not in name:
                mst = law.get("자치법규일련번호") or law.get("MST") or law.get("법령일련번호") or law.get("ID")
                break
        if mst:
            # 2) 조문 본문 → 건폐율 조문에서 용도지역별 퍼센트 파싱
            body = http_get_json("https://www.law.go.kr/DRF/lawService.do", {
                "OC": oc, "target": "ordin", "type": "JSON", "MST": str(mst),
            })
            text = json.dumps(body, ensure_ascii=False)
            zone_key = str(zone_name).replace(" ", "")
            # "제2종일반주거지역 ... 60퍼센트/60%" 형태에서 가장 가까운 퍼센트 추출
            for m in re.finditer(re.escape(zone_key[:7]), text):
                seg = text[m.start(): m.start() + 80]
                pm = re.search(r"(\d{1,3})\s*(?:퍼센트|%)", seg)
                if pm:
                    result = float(pm.group(1))
                    break
    except Exception:
        result = None
    _ORDINANCE_CACHE[cache_key] = result
    return result


def ring_perimeter(points):
    total = 0.0
    for idx in range(len(points)):
        total += segment_length(points[idx], points[(idx + 1) % len(points)])
    return total


def estimate_polygon_width(ring):
    """긴 띠 형태 도로 필지의 대략 폭 ≈ 2*면적/둘레."""
    projected = project_ring(ring)
    area = abs(projected_polygon_area(projected))
    per = ring_perimeter(projected)
    if per <= 0:
        return None
    return max(1.0, 2.0 * area / per)


def jimok_road_width(subject_ring, lat, lon, api_key):
    """인접한 지목='도로'(지번 끝 글자 '도') 필지의 폭을 측정."""
    if not api_key:
        return None
    delta = 0.0004
    best = None
    subject_proj = project_ring(subject_ring)
    subject_segments = polygon_segments(subject_proj)
    for typename in ("lp_pa_cbnd_bubun", "lp_pa_cbnd_bonbun"):
        params = {
            "REQUEST": "GetFeature", "TYPENAME": typename, "VERSION": "1.1.0",
            "MAXFEATURES": "60", "SRSNAME": "EPSG:4326", "OUTPUT": "json",
            "BBOX": f"{lon - delta},{lat - delta},{lon + delta},{lat + delta}", "KEY": api_key,
            "DOMAIN": current_vworld_domain(),
        }
        try:
            data = http_get_json("https://api.vworld.kr/req/wfs", params)
        except Exception:
            continue
        for feature in data.get("features") or []:
            jibun = str((feature.get("properties") or {}).get("jibun") or "")
            if not jibun.endswith("도"):
                continue
            rings = [r for r in geometry_rings(feature.get("geometry") or {}) if len(r) >= 4]
            if not rings:
                continue
            road_ring = max(rings, key=lambda r: abs(projected_polygon_area(project_ring(r))))
            road_proj = [project_point(float(p[0]), float(p[1]), lat) for p in road_ring]
            road_segments = polygon_segments(road_proj)
            min_dist = min(
                (segment_distance(a1, a2, b1, b2) for a1, a2 in subject_segments for b1, b2 in road_segments),
                default=None,
            )
            if min_dist is None or min_dist > 6.0:
                continue
            width = estimate_polygon_width(road_ring)
            if width and (best is None or min_dist < best[1]):
                best = (width, min_dist, jibun)
    if best:
        return {"width": round(best[0], 2), "name": f"지목도로({best[2]})", "source": "지목 도로"}
    return None


def upis_road_width(lat, lon, api_key):
    """도시계획시설(lt_c_upisuq153)에서 인접 계획도로의 폭을 산정."""
    if not api_key:
        return None
    delta = 0.0004
    params = {
        "REQUEST": "GetFeature", "TYPENAME": "lt_c_upisuq153", "VERSION": "1.1.0",
        "MAXFEATURES": "40", "SRSNAME": "EPSG:4326", "OUTPUT": "json",
        "BBOX": f"{lon - delta},{lat - delta},{lon + delta},{lat + delta}", "KEY": api_key,
        "DOMAIN": current_vworld_domain(),
    }
    try:
        data = http_get_json("https://api.vworld.kr/req/wfs", params)
    except Exception:
        return None
    best = None
    for feature in data.get("features") or []:
        props = feature.get("properties") or {}
        names = f"{props.get('lcl_nam','')}{props.get('mls_nam','')}{props.get('dgm_nm','')}"
        if "도로" not in names:
            continue
        area = parse_float(props.get("dgm_ar"))
        length = parse_float(props.get("dgm_lt"))
        width = None
        if area and length and length > 0:
            width = area / length
        else:
            rings = [r for r in geometry_rings(feature.get("geometry") or {}) if len(r) >= 4]
            if rings:
                width = estimate_polygon_width(max(rings, key=lambda r: abs(projected_polygon_area(project_ring(r)))))
        grade = props.get("mls_nam") or props.get("dgm_nm") or "계획도로"
        if width and (best is None or width < best[0]):
            best = (width, grade)
    if best:
        return {"width": round(best[0], 2), "name": f"도시계획도로({best[1]})", "source": "도시계획시설"}
    return None


def boundary_frontage(subject_segments, road_segments, threshold, max_turn_deg=45.0):
    """subject 경계 중 road에 접한 '한 면'의 길이(가장 긴 연속 접면) + 최소거리.
       코너 필지가 한 도로에 두 면 접해도, 방향이 급변(>max_turn)하면 면을 나눠
       주 도로 한 면만 산정한다(여러 면 합산 금지)."""
    n = len(subject_segments)
    if n == 0:
        return 0.0, 1e9
    near = []
    length = []
    bearing = []
    min_dist = None
    for p1, p2 in subject_segments:
        length.append(segment_length(p1, p2))
        bearing.append(math.atan2(p2[1] - p1[1], p2[0] - p1[0]))
        is_near = False
        for r1, r2 in road_segments:
            d = segment_distance(p1, p2, r1, r2)
            min_dist = d if min_dist is None else min(min_dist, d)
            if d <= threshold:
                is_near = True
        near.append(is_near)
    md = min_dist if min_dist is not None else 1e9
    if not any(near):
        return 0.0, md

    max_turn = math.radians(max_turn_deg)

    def ang_diff(a, b):
        d = abs(a - b) % (2 * math.pi)
        return min(d, 2 * math.pi - d)

    # 접면 사이의 경계(비접 세그먼트)에서 시작하도록 회전(원형 wrap 안전 처리)
    start = 0
    if not all(near):
        while near[start]:
            start = (start + 1) % n
    # 도로 방향(라인 방향, 0~pi)으로 정규화: 진행 방향 반대도 같은 면으로 간주
    best = 0.0
    cur = 0.0
    prev_b = None
    for k in range(n):
        i = (start + k) % n
        if near[i]:
            b = bearing[i] % math.pi  # 선분 방향(180° 주기)
            if prev_b is not None and min(abs(b - prev_b), math.pi - abs(b - prev_b)) > max_turn:
                cur = length[i]       # 방향 급변 → 새 면 시작
            else:
                cur += length[i]
            best = max(best, cur)
            prev_b = b
        else:
            cur = 0.0
            prev_b = None
    return best, md


def _wfs_features(typename, lat, lon, api_key, delta=0.0004, maxf=60):
    params = {
        "REQUEST": "GetFeature", "TYPENAME": typename, "VERSION": "1.1.0",
        "MAXFEATURES": str(maxf), "SRSNAME": "EPSG:4326", "OUTPUT": "json",
        "BBOX": f"{lon - delta},{lat - delta},{lon + delta},{lat + delta}",
        "KEY": api_key, "DOMAIN": current_vworld_domain(),
    }
    try:
        return http_get_json("https://api.vworld.kr/req/wfs", params).get("features") or []
    except Exception:
        return []


def jimok_road_features(lat, lon, api_key):
    feats = []
    for typename in ("lp_pa_cbnd_bubun", "lp_pa_cbnd_bonbun"):
        for f in _wfs_features(typename, lat, lon, api_key, 0.0004, 80):
            if str((f.get("properties") or {}).get("jibun") or "").endswith("도"):
                feats.append(f)
    return feats


def upis_road_features(lat, lon, api_key):
    feats = []
    for f in _wfs_features("lt_c_upisuq153", lat, lon, api_key, 0.0004, 40):
        props = f.get("properties") or {}
        names = f"{props.get('lcl_nam', '')}{props.get('mls_nam', '')}{props.get('dgm_nm', '')}"
        if "도로" in names:
            feats.append(f)
    return feats


def resolve_road_width(subject_ring, projected, area, lat, lon, api_key, fallback, slope=1.5):
    """접하는 모든 도로(지목/도시계획/OSM) 각각에 도로사선을 적용해, 최고높이가 '최소'가 되는
       도로를 채택한다(둘 이상 도로 접 시 장변·단변 각각 적용 후 최소값 → 단변 기준 자동 충족).
       채택 도로의 도로폭과 접도길이를 반환."""
    origin_lat = lat
    subj_seg = polygon_segments([project_point(float(p[0]), float(p[1]), origin_lat)
                                 for p in subject_ring if len(p) >= 2])

    def ring_segments(road_ring):
        return polygon_segments([project_point(float(p[0]), float(p[1]), origin_lat) for p in road_ring])

    def best_ring(feature):
        rings = [r for r in geometry_rings(feature.get("geometry") or {}) if len(r) >= 4]
        return max(rings, key=lambda r: abs(projected_polygon_area(project_ring(r)))) if rings else None

    candidates = []

    try:
        for feature in jimok_road_features(lat, lon, api_key):
            road_ring = best_ring(feature)
            if not road_ring:
                continue
            width = estimate_polygon_width(road_ring)
            front, dist = boundary_frontage(subj_seg, ring_segments(road_ring), 4.0)
            if width and front > 0:
                jibun = str((feature.get("properties") or {}).get("jibun") or "")
                candidates.append({"width": round(width, 2), "name": f"지목도로({jibun})",
                                   "source": "지목 도로", "frontage": round(front, 2), "dist": dist})
    except Exception:
        pass

    try:
        for feature in upis_road_features(lat, lon, api_key):
            road_ring = best_ring(feature)
            if not road_ring:
                continue
            props = feature.get("properties") or {}
            ar = parse_float(props.get("dgm_ar"))
            ln = parse_float(props.get("dgm_lt"))
            width = (ar / ln) if (ar and ln and ln > 0) else estimate_polygon_width(road_ring)
            front, dist = boundary_frontage(subj_seg, ring_segments(road_ring), 5.0)
            if width and front > 0:
                grade = props.get("mls_nam") or props.get("dgm_nm") or "계획도로"
                candidates.append({"width": round(width, 2), "name": f"도시계획도로({grade})",
                                   "source": "도시계획시설", "frontage": round(front, 2), "dist": dist})
    except Exception:
        pass

    try:
        for road in osm_roads_near(lat, lon):
            pts = [project_point(c[1], c[0], origin_lat) for c in road["coords"]]
            seg = [(pts[i], pts[i + 1]) for i in range(len(pts) - 1)]
            front, dist = boundary_frontage(subj_seg, seg, 12.0)
            if road.get("width") and front > 0:
                candidates.append({"width": round(road["width"], 2), "name": road.get("name") or "OSM 도로",
                                   "source": "OSM", "frontage": round(front, 2), "dist": dist})
    except Exception:
        pass

    if candidates:
        prio = {"지목 도로": 0, "도시계획시설": 1, "OSM": 2}
        # 각 접도 도로에 도로사선 적용 → 그 도로의 최고높이 = slope×(도로폭 + 깊이),
        # 깊이 = 면적 ÷ 그 도로의 접도길이(장변쪽 도로→작은 깊이=단변). 최고높이가 '최소'인 도로 채택.
        for c in candidates:
            fr = c["frontage"] or math.sqrt(area)
            dep = area / fr
            w_eval = max(c["width"], 4.0) if c["width"] < 4.0 else c["width"]  # 소요너비 미달 4m 보정
            c["maxh"] = slope * (w_eval + dep)
        candidates.sort(key=lambda c: (round(c["maxh"], 3), -c["frontage"], prio.get(c["source"], 9)))
        best = candidates[0]
        detail = " · ".join(
            f"{c['source']} 폭{c['width']}m·접도{c['frontage']}m→최고{round(c['maxh'], 1)}m"
            for c in candidates[:4]
        )
        return best["width"], best["name"], best["source"], detail, best["frontage"]
    return fallback, "접도 도로 자동판정 필요", "기본값", f"기본값={fallback}m", round(math.sqrt(area), 2)


def merge_rings_to_lonlat(member_rings):
    """여러 필지 ring(lon/lat)을 공유 로컬투영에서 union → 합쳐진 외곽 ring(lon/lat) 반환."""
    pts = [p for ring in member_rings for p in ring]
    if not pts:
        return None
    origin_lat = sum(float(p[1]) for p in pts) / len(pts)
    mlon = 111320.0 * max(math.cos(math.radians(origin_lat)), 1e-9)
    mlat = 111320.0
    try:
        from shapely.geometry import Polygon
        from shapely.ops import unary_union
    except Exception:
        return None
    polys = []
    for ring in member_rings:
        coords = [(float(p[0]) * mlon, float(p[1]) * mlat) for p in ring if len(p) >= 2]
        if len(coords) >= 4:
            try:
                poly = Polygon(coords)
                if not poly.is_valid:
                    poly = poly.buffer(0)
                if (not poly.is_empty) and poly.area > 0:
                    polys.append(poly)
            except Exception:
                pass
    if not polys:
        return None
    try:
        merged = unary_union(polys)
    except Exception:
        return None
    if merged.geom_type == "MultiPolygon":
        merged = max(merged.geoms, key=lambda g: g.area)
    try:
        ext = list(merged.exterior.coords)
    except Exception:
        return None
    return [[x / mlon, y / mlat] for (x, y) in ext]


# ── 서울시 지구단위계획 높이계획: 용도지역별 입지(간선부/이면부) 최고높이 상한(m) ──
#   접도조건은 도로폭 20m를 기준으로 간선부/이면부를 구분한다.
def zone_height_cap_pair(zone_name):
    """(간선부, 이면부) 최고높이(m) 튜플. 표에 없는 용도지역은 None."""
    name = str(zone_name or "").replace(" ", "")
    if "제3종일반주거" in name:
        return (60.0, 40.0)
    if "준주거" in name or "준공업" in name:
        return (80.0, 50.0)
    if "근린상업" in name:
        return (90.0, 60.0)
    if "일반상업" in name:
        return (120.0, 70.0)
    return None


# 중심지 위계에 따른 최고높이 조정 가능 범위(±%). 도심·전략지역은 '별도 적용'.
CENTER_RANK_RANGE = {
    "광역중심": 30.0,
    "지역중심": 25.0,
    "지구중심": 20.0,
    "생활권중심": 15.0,
    "그 외": 10.0,
}


def clamp_center_pct(center_rank, requested_pct):
    """위계별 허용 범위로 조정%를 제한. 도심/전략/미지정은 입력값 그대로 통과."""
    rng = CENTER_RANK_RANGE.get(str(center_rank or "").strip())
    pct = requested_pct if requested_pct is not None else 0.0
    if rng is None:
        return pct  # 도심·전략지역(별도 적용) 또는 위계 미지정
    return max(-rng, min(rng, pct))


def fetch_land_characteristics(pnu, api_key, domain=""):
    """VWorld NED 토지특성정보(getLandCharacteristics)에서 PNU의 '최신 기준연도' 레코드를 반환.
    지목·공부면적·용도지역·개별공시지가·이용상황 등 포함. 실패/없음이면 {} (같은 VWorld 키)."""
    pnu = str(pnu or "").strip()
    if not api_key or not pnu:
        return {}
    params = {"key": api_key, "pnu": pnu, "format": "json", "numOfRows": "50", "pageNo": "1"}
    dom = domain or current_vworld_domain()
    if dom:
        params["domain"] = dom
    try:
        data = http_get_json("https://api.vworld.kr/ned/data/getLandCharacteristics", params, retries=2)
    except Exception:
        return {}
    fields = (data.get("landCharacteristicss") or {}).get("field")
    if isinstance(fields, dict):
        fields = [fields]
    if not fields:
        return {}
    # 기준연도(stdrYear)+기준월(stdrMt)이 가장 최신인 레코드
    def key_of(f):
        return (str(f.get("stdrYear") or ""), str(f.get("stdrMt") or ""))
    return max(fields, key=key_of)


def fetch_registered_area(pnu, api_key, domain=""):
    """공부면적(토지면적, lndpclAr, ㎡) — 최신 토지특성 레코드 기준. 없으면 None."""
    rec = fetch_land_characteristics(pnu, api_key, domain)
    ar = parse_float(rec.get("lndpclAr")) if rec else None
    return ar if (ar and ar > 0) else None


def fetch_land_ladfrl(pnu, api_key, domain=""):
    """VWorld NED 토지임야(ladfrlList): 소유구분(posesnSeCodeNm)·공유인수(cnrsPsnCo) 등 반환.
    소유자 '이름'은 제공하지 않고 구분만 제공. 실패 시 {}."""
    pnu = str(pnu or "").strip()
    if not api_key or not pnu:
        return {}
    params = {"key": api_key, "pnu": pnu, "format": "json", "numOfRows": "5", "pageNo": "1"}
    dom = domain or current_vworld_domain()
    if dom:
        params["domain"] = dom
    try:
        data = http_get_json("https://api.vworld.kr/ned/data/ladfrlList", params, retries=2)
    except Exception:
        return {}
    vo = (data.get("ladfrlVOList") or {}).get("ladfrlVOList")
    if isinstance(vo, dict):
        vo = [vo]
    return vo[0] if vo else {}


def simplify_owner_type(name):
    """소유구분명을 개인/법인/국공유지/외국인 등으로 단순화."""
    nm = str(name or "").strip()
    if not nm:
        return ""
    if "법인" in nm:
        return "법인"
    if "외국" in nm:
        return "외국인"
    if "개인" in nm:
        return "개인"
    if any(t in nm for t in ("국유", "공유", "국가", "도유", "시유", "군유", "구유", "공공", "지자체")):
        return "국공유지"
    if "종중" in nm:
        return "종중"
    return nm


def analyze_selected_parcels(payload):
    api_key = str(payload.get("apiKey") or "").strip()
    if not api_key:
        raise ValueError("선택필지 자동분석에는 VWorld API 키가 필요합니다.")
    parcel_items = payload.get("parcelListItems") or []
    if not parcel_items:
        raise ValueError("지도에서 필지를 먼저 선택해 주세요.")
    slope_multiplier = parse_float(payload.get("slopeMultiplier"), 1.5)
    fallback_road_width = 8.0
    review_note = str(payload.get("reviewNote") or "").strip()
    law_oc = str(payload.get("lawApiKey") or "").strip()
    # 수동 입력값(있으면 우선)
    manual_district = parse_float(payload.get("districtPlanCoverage"), None)
    manual_ordinance = parse_float(payload.get("ordinanceCoverage"), None)
    manual_front = parse_float(payload.get("setbackFront"), parse_float(payload.get("setback"), None))  # 전역 기본 전면건축선후퇴(m)
    manual_rear = parse_float(payload.get("setbackRear"), None)  # 전역 기본 후면건축선후퇴(m)
    # 획지(group)별 건축선후퇴: { "<group>": {"front": x, "rear": y} }
    lot_setbacks = payload.get("lotSetbacks") or {}
    # 획지(group)별 용도지역 상향: { "<group>": "준주거지역" }
    lot_zones = payload.get("lotZoneOverride") or {}
    # 중심지 위계 조정(①): 위계 + 조정%
    center_rank = str(payload.get("centerRank") or "").strip()
    center_pct = clamp_center_pct(center_rank, parse_float(payload.get("centerAdjustPct"), 0.0))
    # 공부면적(토지특성) 자동조회 사용 여부(기본 사용; 실패 시 도형면적으로 폴백)
    use_reg_area = payload.get("useRegArea", True) is not False
    rows = []
    total_floor_area = 0.0
    total_volume = 0.0

    # group 값으로 획지 묶음 구성(없으면 각 필지가 단독 획지). 첫 등장 순서 유지.
    grouped = []
    group_keys = []
    index_of = {}
    for item in parcel_items:
        g = item.get("group")
        key = g if g is not None else f"_solo_{len(grouped)}"
        if key not in index_of:
            index_of[key] = len(grouped)
            grouped.append([])
            group_keys.append(g)
        grouped[index_of[key]].append(item)

    for grp_idx, members in enumerate(grouped):
        grp_key = group_keys[grp_idx]
        member_rings = []
        member_parcels = []
        member_area_sum = 0.0
        member_reg_sum = 0.0
        all_have_reg = True   # 모든 멤버에 공부면적이 조회됐는지(획지 합산용)
        for item in members:
            m_lat = float(item.get("lat"))
            m_lon = float(item.get("lon"))
            m_parcel = str(item.get("parcel") or "").strip() or reverse_geocode_parcel(m_lat, m_lon, api_key)
            feature = get_parcel_feature(m_lat, m_lon, api_key)["features"][0]
            ring = primary_ring(feature)
            member_rings.append(ring)
            member_parcels.append(m_parcel)
            # 도형면적: 멀티폴리곤·구멍을 정확히 반영(외곽−구멍 합)
            member_area_sum += feature_geom_area(feature)
            # 공부면적(토지특성, lndpclAr) — PNU로 자동조회(같은 VWorld 키)
            reg = None
            if use_reg_area:
                pnu = str((feature.get("properties") or {}).get("pnu") or "").strip()
                reg = fetch_registered_area(pnu, api_key)
            if reg and reg > 0:
                member_reg_sum += reg
            else:
                all_have_reg = False

        geom_area = member_area_sum
        if len(member_rings) <= 1:
            ring = member_rings[0]
        else:
            # 여러 필지를 하나의 획지로 합침(외곽선은 union, 면적은 비중첩 합)
            merged_ring = merge_rings_to_lonlat(member_rings)
            ring = merged_ring if merged_ring else member_rings[0]
        reg_area = member_reg_sum if (use_reg_area and all_have_reg and member_reg_sum > 0) else None
        # 대지면적(산정 기준): 공부면적 있으면 공부, 없으면 도형
        area = reg_area if reg_area else geom_area
        area_source = "공부면적(토지특성)" if reg_area else "도형면적(연속지적)"
        projected = project_ring(ring)
        c_lat, c_lon = lonlat_centroid(ring)

        # 용도지역 자동판별(획지 중심)
        landuse = get_landuse(c_lat, c_lon, api_key) or {}
        zone_auto = landuse.get("zone")
        sigungu = landuse.get("sigungu")

        # 획지별 용도지역 상향(지구단위계획 등): 지정 시 건폐율·최고높이 상한을 상향 용도로 산정
        zone_ov = str(lot_zones.get(str(grp_key)) or lot_zones.get(grp_key) or "").strip()
        zone_upgraded = bool(zone_ov and zone_ov != (zone_auto or ""))
        zone = zone_ov if zone_ov else zone_auto

        # 건폐율: 지구단위계획(입력) > 조례(입력/법제처) > 시행령 상한 > 없음
        coverage = None
        cov_source = "확인 필요"
        if manual_district:
            coverage, cov_source = manual_district, "지구단위계획(입력)"
        elif manual_ordinance:
            coverage, cov_source = manual_ordinance, "조례(입력)"
        else:
            ord_auto = fetch_ordinance_coverage(sigungu, zone, law_oc) if law_oc else None
            if ord_auto:
                coverage, cov_source = ord_auto, f"조례(법제처·{zone})"
            else:
                cap = coverage_cap_for(zone)
                if cap:
                    coverage, cov_source = cap, f"시행령 상한({zone})"
        if not coverage:
            coverage, cov_source = 60.0, "확인 필요(기본 60% 가정)"

        # 도로폭: 경계에 가장 많이 접한 도로 기준(지목도로/도시계획/OSM 중 frontage 최대)
        road_width, road_name, road_source, road_detail, frontage = resolve_road_width(
            ring, projected, area, c_lat, c_lon, api_key, fallback_road_width, slope_multiplier
        )
        if not frontage:
            frontage = math.sqrt(area)
        # 깊이는 채택 도로의 접도길이 기준
        depth = area / frontage
        coverage_ratio = coverage / 100.0 if coverage > 1 else coverage

        # 건축선 반영(서울시 높이계획 방법론):
        #  - 소요너비 미달 도로(<4m)는 도로폭 4m로 보정
        #  - 전면건축선후퇴는 도로사선 기준거리에 '가산': 최저=1.5×(도로폭+전면후퇴),
        #    최고=1.5×(도로폭+전면후퇴+(획지깊이-후면후퇴))
        #  - 바닥면적은 대지면적 '전체'×건폐율(후퇴분을 면적에서 빼지 않음)
        raw_road = road_width
        road_width_used = max(raw_road, 4.0) if raw_road < 4.0 else raw_road
        # 건축선후퇴: 획지(group)별 입력값 우선, 없으면 전역 기본값
        lot_sb = lot_setbacks.get(str(grp_key)) or lot_setbacks.get(grp_key) or {}
        lot_front = parse_float(lot_sb.get("front"), None)
        lot_rear = parse_float(lot_sb.get("rear"), None)
        eff_front = lot_front if lot_front is not None else manual_front
        eff_rear = lot_rear if lot_rear is not None else manual_rear
        front_setback = eff_front if (eff_front is not None and eff_front > 0) else 0.0
        rear_setback = eff_rear if (eff_rear is not None and eff_rear > 0) else 0.0
        eff_depth = max(depth - rear_setback, 0.0)
        sine_base = road_width_used + front_setback

        floor_area = area * coverage_ratio
        min_height = slope_multiplier * sine_base
        max_height = slope_multiplier * (sine_base + eff_depth)
        avg_height = (min_height + max_height) / 2.0
        volume = floor_area * avg_height

        # 입지(간선부/이면부) 구분 + 용도지역별 최고높이 상한(②) → 중심지 위계 조정(①)
        is_arterial = road_width_used >= 20.0
        arterial_class = "간선부" if is_arterial else "이면부"
        cap_pair = zone_height_cap_pair(zone)
        height_cap_base = (cap_pair[0] if is_arterial else cap_pair[1]) if cap_pair else None
        if height_cap_base is not None:
            height_cap_adjusted = round(height_cap_base * (1.0 + center_pct / 100.0), 2)
            cap_exceeded = avg_height > height_cap_adjusted + 1e-6
        else:
            height_cap_adjusted = None
            cap_exceeded = False

        if front_setback or rear_setback or road_width_used != raw_road:
            road_detail += (f" · 건축선후퇴 전면{round(front_setback, 2)}m"
                            + (f"·후면{round(rear_setback, 2)}m" if rear_setback else "")
                            + (f", 도로폭 {round(road_width_used, 1)}m 적용" if road_width_used != raw_road else ""))
        total_floor_area += floor_area
        total_volume += volume
        cov_ok = cov_source.startswith(("지구단위", "조례", "시행령"))
        confidence = "상" if (road_source in ("지목 도로", "도시계획시설") and cov_ok) else "중" if cov_ok else "하"

        is_hoekji = len(member_parcels) > 1
        label = f"{member_parcels[0]} 외 {len(member_parcels) - 1}필지" if is_hoekji else member_parcels[0]
        rows.append(
            {
                "parcel": label,
                "parcelCount": len(member_parcels),
                "members": ", ".join(member_parcels),
                "isHoekji": is_hoekji,
                "zone": zone or "-",
                "zoneAuto": zone_auto or "-",
                "zoneUpgraded": zone_upgraded,
                "area": round(area, 3),
                "geomArea": round(geom_area, 3),
                "regArea": round(reg_area, 3) if reg_area else None,
                "areaSource": area_source,
                "frontSetback": round(front_setback, 3),
                "rearSetback": round(rear_setback, 3),
                "coverageRatio": round(coverage_ratio * 100, 3),
                "coverageSource": cov_source,
                "floorArea": round(floor_area, 3),
                "roadName": road_name,
                "roadWidth": round(road_width_used, 3),
                "roadWidthRaw": round(raw_road, 3),
                "roadSource": road_source,
                "roadDetail": road_detail,
                "frontage": round(frontage, 3),
                "depth": round(depth, 3),
                "minHeight": round(min_height, 3),
                "maxHeight": round(max_height, 3),
                "avgHeight": round(avg_height, 3),
                "volume": round(volume, 3),
                "arterialClass": arterial_class,
                "heightCapBase": height_cap_base,
                "heightCapAdjusted": height_cap_adjusted,
                "centerRank": center_rank or "-",
                "centerAdjustPct": round(center_pct, 1),
                "capExceeded": cap_exceeded,
                "method": "획지 자동분석" if is_hoekji else "선택필지 자동분석",
                "confidence": confidence,
                "note": review_note,
            }
        )
    if total_floor_area <= 0:
        raise ValueError("건폐율 적용 바닥면적 합계가 0입니다.")
    return {
        "rows": rows,
        "totals": {
            "parcelCount": len(rows),
            "floorArea": round(total_floor_area, 3),
            "volume": round(total_volume, 3),
            "blockAverageHeight": round(total_volume / total_floor_area, 3),
            "capExceededCount": sum(1 for rrow in rows if rrow.get("capExceeded")),
            "centerRank": center_rank or "-",
            "centerAdjustPct": round(center_pct, 1),
        },
    }


def facility_category(tags):
    railway = tags.get("railway")
    station = tags.get("station")
    amenity = tags.get("amenity")
    highway = tags.get("highway")
    leisure = tags.get("leisure")
    tourism = tags.get("tourism")
    office = tags.get("office")
    building = tags.get("building")
    landuse = tags.get("landuse")
    name = tags.get("name") or tags.get("name:ko") or ""

    if railway in {"station", "halt", "subway_entrance"} or station == "subway":
        return "교통"
    if highway == "bus_station" or amenity == "bus_station":
        return "교통"
    if amenity in {"townhall", "courthouse", "police", "fire_station", "post_office"} or office == "government":
        return "공공"
    if amenity == "kindergarten":
        return None
    if amenity in {"school", "university", "college"}:
        return "교육"
    if amenity == "hospital":
        return "의료"
    if building in {"apartments", "residential"} or landuse == "residential":
        return "주거"
    if "아파트" in name or "주공" in name or "래미안" in name or "자이" in name:
        return "주거"
    if leisure in {"stadium", "sports_centre", "park"} or tourism in {"attraction", "museum"}:
        return "생활"
    if "시청" in name or "구청" in name or "경찰" in name or "소방" in name:
        return "공공"
    if "역" in name or "정류장" in name:
        return "교통"
    return None


def allowed_facility(name, category):
    if category == "교육" and "유치원" in name:
        return False
    return True


def get_context_facilities(lat, lon, radius=2500):
    curated = curated_facilities(lat, lon, radius)
    query = f"""
    [out:json][timeout:20];
    (
      node(around:{int(radius)},{lat},{lon})["railway"~"station|halt|subway_entrance"];
      node(around:{int(radius)},{lat},{lon})["station"="subway"];
      node(around:{int(radius)},{lat},{lon})["highway"="bus_station"];
      node(around:{int(radius)},{lat},{lon})["amenity"~"bus_station|townhall|courthouse|police|fire_station|post_office|school|university|college|hospital"];
      node(around:{int(radius)},{lat},{lon})["leisure"~"stadium|sports_centre|park"];
      way(around:{int(radius)},{lat},{lon})["amenity"~"townhall|courthouse|police|fire_station|school|university|hospital"];
      way(around:{int(radius)},{lat},{lon})["leisure"~"stadium|sports_centre|park"];
      way(around:{int(radius)},{lat},{lon})["building"~"apartments|residential"];
      way(around:{int(radius)},{lat},{lon})["landuse"="residential"];
      relation(around:{int(radius)},{lat},{lon})["amenity"~"townhall|courthouse|police|fire_station|school|university|hospital"];
      relation(around:{int(radius)},{lat},{lon})["leisure"~"stadium|sports_centre|park"];
      relation(around:{int(radius)},{lat},{lon})["building"~"apartments|residential"];
      relation(around:{int(radius)},{lat},{lon})["landuse"="residential"];
    );
    out center tags 160;
    """
    req = urllib.request.Request(
        OVERPASS_URL,
        data=urllib.parse.urlencode({"data": query}).encode("utf-8"),
        headers={"User-Agent": "CadastreMapBridge/1.0"},
    )
    try:
        with urllib.request.urlopen(req, timeout=30) as res:
            data = json.loads(res.read().decode("utf-8", errors="replace"))
    except Exception:
        return curated

    facilities = list(curated)
    seen = set()
    for item in facilities:
        seen.add((item["name"], round(float(item["lat"]), 5), round(float(item["lon"]), 5)))
    for element in data.get("elements", []):
        tags = element.get("tags", {})
        name = tags.get("name:ko") or tags.get("name")
        if not name:
            continue
        item_lat = element.get("lat") or element.get("center", {}).get("lat")
        item_lon = element.get("lon") or element.get("center", {}).get("lon")
        if item_lat is None or item_lon is None:
            continue
        category = facility_category(tags)
        if not category:
            continue
        if not allowed_facility(name, category):
            continue
        key = (name, round(float(item_lat), 5), round(float(item_lon), 5))
        if key in seen:
            continue
        seen.add(key)
        distance = haversine_m(lat, lon, float(item_lat), float(item_lon))
        facilities.append(
            {
                "name": name,
                "category": category,
                "lat": float(item_lat),
                "lon": float(item_lon),
                "distance": round(distance),
            }
        )

    facilities.sort(key=lambda item: (category_rank(item["category"]), item["distance"]))
    facilities = dedupe_facilities(facilities)
    caps = {"교통": 10, "공공": 8, "생활": 8, "교육": 8, "의료": 6, "주거": 18}
    counts = {}
    selected = []
    for item in facilities:
        category = item["category"]
        counts[category] = counts.get(category, 0)
        if counts[category] >= caps.get(category, 3):
            continue
        selected.append(item)
        counts[category] += 1
        if len(selected) >= 58:
            break
    return selected


def canonical_facility_name(name, category):
    cleaned = clean_facility_name(name, category)
    if category == "교통":
        if "정류장" not in cleaned and "터미널" not in cleaned and not cleaned.endswith("역"):
            cleaned = f"{cleaned}역"
        cleaned = re.sub(r"(역|驛)$", "역", cleaned)
        cleaned = re.sub(r"(역)(역)$", "역", cleaned)
    return cleaned


def dedupe_facilities(facilities):
    result = []
    seen = set()
    for item in facilities:
        canonical = canonical_facility_name(item["name"], item["category"])
        if not canonical:
            continue
        key = (item["category"], canonical)
        if key in seen:
            continue
        item = dict(item)
        item["name"] = canonical
        seen.add(key)
        result.append(item)
    return result


def curated_facilities(lat, lon, radius):
    candidates = [
        ("안양종합운동장", "생활", 37.4050, 126.9509),
        ("안양시청", "공공", 37.3944, 126.9568),
        ("동안구청", "공공", 37.3926, 126.9518),
        ("범계역", "교통", 37.3898, 126.9508),
        ("평촌역", "교통", 37.3943, 126.9638),
        ("인덕원역", "교통", 37.4019, 126.9767),
        ("안양역", "교통", 37.4019, 126.9227),
        ("평촌중앙공원", "생활", 37.3917, 126.9575),
        ("한림대학교성심병원", "의료", 37.3913, 126.9620),
        ("안양동안경찰서", "공공", 37.3948, 126.9589),
        ("삼호아파트", "주거", 37.4008, 126.9488),
        ("비산삼성래미안", "주거", 37.4073, 126.9468),
        ("관악성원아파트", "주거", 37.3991, 126.9568),
        ("평촌래미안푸르지오", "주거", 37.3916, 126.9487),
        ("샛별한양아파트", "주거", 37.3884, 126.9555),
    ]
    facilities = []
    for name, category, item_lat, item_lon in candidates:
        distance = haversine_m(lat, lon, item_lat, item_lon)
        if distance <= max(radius, 3500):
            facilities.append(
                {
                    "name": name,
                    "category": category,
                    "lat": item_lat,
                    "lon": item_lon,
                    "distance": round(distance),
                }
            )
    return facilities


def haversine_m(lat1, lon1, lat2, lon2):
    earth = 6371000
    phi1 = math.radians(lat1)
    phi2 = math.radians(lat2)
    d_phi = math.radians(lat2 - lat1)
    d_lambda = math.radians(lon2 - lon1)
    a = math.sin(d_phi / 2) ** 2 + math.cos(phi1) * math.cos(phi2) * math.sin(d_lambda / 2) ** 2
    return earth * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))


def category_rank(category):
    ranks = {"교통": 0, "공공": 1, "생활": 2, "주거": 3, "교육": 4, "의료": 5}
    return ranks.get(category, 9)


def latlon_to_pixel(lat, lon, zoom):
    sin_lat = math.sin(math.radians(lat))
    world = 256 * (2**zoom)
    x = (lon + 180.0) / 360.0 * world
    y = (0.5 - math.log((1 + sin_lat) / (1 - sin_lat)) / (4 * math.pi)) * world
    return x, y


def tile_image(z, x, y, map_type="satellite"):
    templates = {
        "satellite": ESRI_TILE_URL,
        "street": CARTO_VOYAGER_TILE_URL,
        "blank": CARTO_LIGHT_NOLABEL_TILE_URL,
    }
    template = templates.get(map_type, ESRI_TILE_URL)
    url = template.format(z=z, x=x, y=y)
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=30) as res:
        return Image.open(res).convert("RGB")


def load_font(size, bold=False):
    candidates = [
        Path("C:/Windows/Fonts/malgunbd.ttf" if bold else "C:/Windows/Fonts/malgun.ttf"),
        Path("C:/Windows/Fonts/arialbd.ttf" if bold else "C:/Windows/Fonts/arial.ttf"),
        # Linux(컨테이너) 한글 폰트 — Dockerfile에서 fonts-nanum 설치
        Path("/usr/share/fonts/truetype/nanum/NanumGothicBold.ttf" if bold else "/usr/share/fonts/truetype/nanum/NanumGothic.ttf"),
        Path("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf" if bold else "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
    ]
    for candidate in candidates:
        if candidate.exists():
            return ImageFont.truetype(str(candidate), size)
    return ImageFont.load_default()


def draw_centered_text(draw, box, text, font, fill):
    left, top, right, bottom = box
    bbox = draw.textbbox((0, 0), text, font=font)
    width = bbox[2] - bbox[0]
    height = bbox[3] - bbox[1]
    draw.text((left + (right - left - width) / 2, top + (bottom - top - height) / 2), text, font=font, fill=fill)


def choose_zoom(lat, radius, span):
    for zoom in range(18, 10, -1):
        meters_per_pixel = 156543.03392 * math.cos(math.radians(lat)) / (2**zoom)
        if radius / meters_per_pixel <= span * 0.47:
            return zoom
    return 11


def draw_text_with_outline(draw, xy, text, font, fill, outline=(20, 20, 20, 255), width=2):
    x, y = xy
    for dx in range(-width, width + 1):
        for dy in range(-width, width + 1):
            if dx or dy:
                draw.text((x + dx, y + dy), text, font=font, fill=outline)
    draw.text((x, y), text, font=font, fill=fill)


def clean_facility_name(name, category):
    name = " ".join(str(name or "").replace("\n", " ").split())
    name = re.sub(r"\s+\d{1,4}$", "", name)
    name = re.sub(r"\s+\d{1,4}(동|호)$", "", name)
    if category == "주거":
        name = re.sub(r"\s+[가-힣A-Za-z0-9]{1,4}동$", "", name)
        name = re.sub(r"\s+[가-힣A-Za-z0-9]{1,4}호$", "", name)
        name = re.sub(r"\s*\d{1,4}$", "", name)
    return name.strip()


def facility_label_box(draw, x, y, facility, font):
    text = clean_facility_name(facility["name"], facility["category"])
    if not text or text.isdigit() or len(text) < 2:
        return None
    bbox = draw.textbbox((0, 0), text, font=font)
    text_w = bbox[2] - bbox[0]
    text_h = bbox[3] - bbox[1]
    return (x - text_w / 2 - 4, y - text_h / 2 - 3, x + text_w / 2 + 4, y + text_h / 2 + 3, text)


def boxes_overlap(a, b, padding=3):
    return not (a[2] + padding < b[0] or b[2] + padding < a[0] or a[3] + padding < b[1] or b[3] + padding < a[1])


def box_center(box):
    return ((box[0] + box[2]) / 2, (box[1] + box[3]) / 2)


def box_distance(a, b):
    ax, ay = box_center(a)
    bx, by = box_center(b)
    return math.hypot(ax - bx, ay - by)


def crowded_label_count(box, placed_boxes, distance):
    return sum(1 for placed in placed_boxes if box_distance(box, placed) < distance)


def draw_facility_label(draw, box, text, facility, font):
    colors = {
        "교통": (69, 163, 255, 255),
        "공공": (216, 117, 255, 255),
        "생활": (53, 232, 111, 255),
        "교육": (255, 212, 0, 255),
        "의료": (255, 87, 87, 255),
        "주거": (255, 184, 74, 255),
    }
    color = colors.get(facility["category"], (255, 255, 255, 255))
    draw_text_with_outline(draw, (box[0] + 4, box[1] + 3), text, font, color, outline=(0, 0, 0, 255), width=2)


def draw_dashed_ellipse(draw, box, fill, width=2, dash_degrees=1.2, gap_degrees=1.2):
    start = 0
    while start < 360:
        end = min(start + dash_degrees, 360)
        draw.arc(box, start=start, end=end, fill=fill, width=width)
        start += dash_degrees + gap_degrees


def format_radius_label(radius_m):
    if radius_m >= 1000:
        km = radius_m / 1000
        if abs(km - round(km)) < 0.05:
            return f"{int(round(km))}km"
        return f"{km:.1f}km"
    return f"{int(round(radius_m))}m"


def radius_ring_values(radius_m):
    if radius_m <= 1000:
        values = [value for value in (200, 500, 1000) if value <= radius_m]
    else:
        values = [value for value in (500, 1000, 3000, 5000, 10000, 25000) if value <= radius_m]
    return sorted(values)


def draw_radius_ring(draw, cx, cy, pixel_radius, label, font, canvas_size, scale=1.0, primary=False):
    line_width = max(2, round((2 if primary else 1) * scale))
    outline_width = max(line_width + 2, round((4 if primary else 2) * scale))
    dash = 1.2 if primary else 1.0
    gap = 1.2 if primary else 1.0
    box = (cx - pixel_radius, cy - pixel_radius, cx + pixel_radius, cy + pixel_radius)
    # 흰 헤일로(어두운 위성 위 가독성) + 블루 선(UI 기본색 #1b4f9c와 통일)
    draw_dashed_ellipse(draw, box, (255, 255, 255, 235), width=outline_width, dash_degrees=dash, gap_degrees=gap)
    draw_dashed_ellipse(draw, box, (27, 79, 156, 255), width=line_width, dash_degrees=dash, gap_degrees=gap)

    bbox = draw.textbbox((0, 0), label, font=font)
    label_w = bbox[2] - bbox[0]
    label_h = bbox[3] - bbox[1]
    angle = math.radians(-28)
    label_x = cx + math.cos(angle) * pixel_radius - label_w / 2
    label_y = cy + math.sin(angle) * pixel_radius - label_h / 2
    canvas_w, canvas_h = canvas_size
    label_x = max(12, min(canvas_w - label_w - 12, label_x))
    label_y = max(12, min(canvas_h - label_h - 12, label_y))
    draw_text_with_outline(draw, (label_x, label_y), label, font, (255, 255, 255, 255), outline=(20, 20, 20, 255), width=2)


def make_satellite_map(address, api_key, width=3200, height=2000, radius=3000, show_facilities=True, categories=None, map_type="satellite", show_radius=True):
    lat, lon, refined = geocode(address, api_key)
    zoom = choose_zoom(lat, radius, min(width, height))
    center_px, center_py = latlon_to_pixel(lat, lon, zoom)
    left_px = center_px - width / 2
    top_px = center_py - height / 2
    first_tile_x = math.floor(left_px / 256)
    first_tile_y = math.floor(top_px / 256)
    last_tile_x = math.floor((left_px + width) / 256)
    last_tile_y = math.floor((top_px + height) / 256)

    blank_mode = map_type == "blank"
    background = (248, 248, 246) if blank_mode else (20, 22, 20)
    canvas = Image.new("RGB", ((last_tile_x - first_tile_x + 1) * 256, (last_tile_y - first_tile_y + 1) * 256), background)
    max_tile = 2**zoom
    for ty in range(first_tile_y, last_tile_y + 1):
        for tx in range(first_tile_x, last_tile_x + 1):
            try:
                tile = tile_image(zoom, tx % max_tile, ty, map_type=map_type)
                canvas.paste(tile, ((tx - first_tile_x) * 256, (ty - first_tile_y) * 256))
            except Exception:
                pass

    crop_left = int(round(left_px - first_tile_x * 256))
    crop_top = int(round(top_px - first_tile_y * 256))
    img = canvas.crop((crop_left, crop_top, crop_left + width, crop_top + height)).convert("RGBA")
    draw = ImageDraw.Draw(img)

    scale = min(width / 1600, height / 1000)
    label_font = load_font(max(18, round(17 * scale)), bold=True)
    small_font = load_font(16)
    facility_font = load_font(max(13, round(10 * scale)), bold=True)

    cx, cy = width // 2, height // 2
    meters_per_pixel = 156543.03392 * math.cos(math.radians(lat)) / (2**zoom)
    pixel_radius = max(32, min(min(width, height) * 0.47, radius / meters_per_pixel))

    if show_radius:
        ring_values = radius_ring_values(radius)
        for ring_radius_m in ring_values:
            ring_pixel_radius = max(24, ring_radius_m / meters_per_pixel)
            ring_label = format_radius_label(ring_radius_m)
            draw_radius_ring(
                draw,
                cx,
                cy,
                ring_pixel_radius,
                ring_label,
                label_font,
                (width, height),
                scale=scale,
                primary=bool(ring_values and ring_radius_m == ring_values[-1]),
            )
    dot_r = max(11, round(11 * scale))
    dot_w = max(4, round(4 * scale))
    draw.ellipse((cx - dot_r, cy - dot_r, cx + dot_r, cy + dot_r), fill=(27, 79, 156, 240), outline=(255, 255, 255, 255), width=dot_w)

    if show_facilities:
        try:
            facilities = get_context_facilities(lat, lon, radius)
        except Exception:
            facilities = curated_facilities(lat, lon, radius)
        if categories:
            facilities = [facility for facility in facilities if facility["category"] in categories]
        placed_boxes = []
        placed_names = set()
        facilities = sorted(facilities, key=lambda item: (category_rank(item["category"]), item["distance"]))
        offset_scale = max(0.85, scale * 0.72)
        offsets = [
            (0, 0), (0, -36), (0, 36), (52, 0), (-52, 0),
            (58, -36), (-58, -36), (58, 36), (-58, 36),
            (0, -72), (0, 72), (104, 0), (-104, 0),
            (104, -58), (-104, -58), (104, 58), (-104, 58),
            (0, -110), (0, 110), (150, 0), (-150, 0),
        ]
        offsets = [(dx * offset_scale, dy * offset_scale) for dx, dy in offsets]
        for facility in facilities:
            fx, fy = latlon_to_pixel(facility["lat"], facility["lon"], zoom)
            sx = cx + (fx - center_px)
            sy = cy + (fy - center_py)
            if not (40 < sx < width - 40 and 40 < sy < height - 40):
                continue
            for dx, dy in offsets:
                box = facility_label_box(draw, sx + dx, sy + dy, facility, facility_font)
                if not box:
                    break
                if box[4] in placed_names:
                    break
                if box[0] < 8 or box[1] < 8 or box[2] > width - 8 or box[3] > height - 32:
                    continue
                if any(boxes_overlap(box, placed, padding=max(4, round(5 * scale))) for placed in placed_boxes):
                    continue
                if crowded_label_count(box, placed_boxes, distance=max(78, round(54 * scale))) >= 2:
                    continue
                draw_facility_label(draw, box, box[4], facility, facility_font)
                placed_boxes.append(box)
                placed_names.add(box[4])
                break

    footer = {
        "satellite": "Satellite imagery: Esri | Geocoding: VWorld",
        "street": "Map: CARTO (OpenStreetMap data) | Geocoding: VWorld",
        "blank": "Blank base map | Geocoding: VWorld",
    }.get(map_type, "Geocoding: VWorld")
    bbox = draw.textbbox((0, 0), footer, font=small_font)
    draw.rectangle((width - (bbox[2] - bbox[0]) - 22, height - 30, width, height), fill=(255, 255, 255, 210))
    draw.text((width - (bbox[2] - bbox[0]) - 12, height - 26), footer, font=small_font, fill=(40, 40, 40, 255))

    out = ROOT / f"location_map_{time.strftime('%Y%m%d_%H%M%S')}.png"
    img.convert("RGB").save(out, "PNG")
    return out, lat, lon, refined


def select_folder(initial_dir=""):
    initial = str(initial_dir or DEFAULT_OUT_DIR)
    script = r"""
& {
  param([string]$InitialPath)
  Add-Type -AssemblyName System.Windows.Forms
  Add-Type -AssemblyName System.Drawing
  [Console]::OutputEncoding = [System.Text.Encoding]::UTF8
  $owner = New-Object System.Windows.Forms.Form
  $owner.StartPosition = "Manual"
  $owner.Size = New-Object System.Drawing.Size(1, 1)
  $owner.Location = New-Object System.Drawing.Point(-32000, -32000)
  $owner.ShowInTaskbar = $false
  $owner.TopMost = $true
  $owner.Show()
  $owner.Activate()
  $dialog = New-Object System.Windows.Forms.FolderBrowserDialog
  $dialog.Description = "연속지적 저장 폴더 선택"
  $dialog.ShowNewFolderButton = $true
  if ($InitialPath -and (Test-Path -LiteralPath $InitialPath)) {
    $dialog.SelectedPath = $InitialPath
  }
  try {
    if ($dialog.ShowDialog($owner) -eq [System.Windows.Forms.DialogResult]::OK) {
      Write-Output $dialog.SelectedPath
    }
  } finally {
    $owner.Close()
    $owner.Dispose()
  }
}
"""
    creationflags = getattr(subprocess, "CREATE_NO_WINDOW", 0)
    result = subprocess.run(
        [
            "powershell.exe",
            "-NoProfile",
            "-STA",
            "-ExecutionPolicy",
            "Bypass",
            "-Command",
            script,
            initial,
        ],
        cwd=str(ROOT),
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
        creationflags=creationflags,
        timeout=600,
    )
    if result.returncode != 0:
        raise RuntimeError((result.stderr or result.stdout or "폴더 선택창을 열지 못했습니다.").strip())
    return result.stdout.strip()


def save_parcel_list_upload(payload, out_dir):
    paths = []
    upload = payload.get("parcelListFile") or {}
    name = str(upload.get("name") or "").strip()
    data_url = str(upload.get("data") or "")
    upload_dir = out_dir / "parcel_list_uploads"
    upload_dir.mkdir(parents=True, exist_ok=True)
    if name and data_url:
        if "," in data_url:
            data_url = data_url.split(",", 1)[1]
        suffix = Path(name).suffix.lower()
        if suffix not in {".xlsx", ".csv", ".txt"}:
            raise ValueError("지번 리스트는 .xlsx, .csv, .txt 파일만 첨부할 수 있습니다.")
        safe_name = re.sub(r'[\\/:*?"<>|]+', "_", Path(name).name)
        target = upload_dir / f"{time.strftime('%Y%m%d_%H%M%S')}_{safe_name}"
        target.write_bytes(base64.b64decode(data_url))
        paths.append(target)

    parcel_items = payload.get("parcelListItems") or []
    if parcel_items:
        target = upload_dir / f"{time.strftime('%Y%m%d_%H%M%S')}_map_selected_parcels.csv"
        lines = ["parcel"]
        for item in parcel_items:
            text = str(item.get("parcel") if isinstance(item, dict) else item).strip()
            if text:
                lines.append('"' + text.replace('"', '""') + '"')
        if len(lines) > 1:
            target.write_text("\n".join(lines), encoding="utf-8-sig")
            paths.append(target)
    return paths


def normalize_header(value):
    return re.sub(r"[^0-9a-zA-Z가-힣]+", "", str(value or "")).lower()


def parse_float(value, default=None):
    if value is None:
        return default
    text = str(value).strip()
    if not text:
        return default
    text = text.replace(",", "").replace("㎡", "").replace("m2", "").replace("m", "").replace("%", "")
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return default
    return float(match.group(0))


def row_value(row, aliases, default=None):
    for alias in aliases:
        key = normalize_header(alias)
        if key in row and str(row[key]).strip():
            return row[key]
    return default


# 표 헤더 자동 인식용 키워드(필지 조서·평균높이 양쪽 열 포함).
KNOWN_HEADER_KEYS = {
    "주소", "소재지", "소재지지번", "소재지번", "소재", "위치", "address", "addr",
    "지번", "번지", "필지", "필지명", "parcel", "lot", "pnu",
    "지목", "소유구분", "용도지역",
    "면적", "대지면적", "필지면적", "공부면적", "도형면적", "sitearea", "lotarea", "area",
    "건폐율", "coverage", "coverageratio", "도로폭", "도로너비", "접도폭", "도로폭원", "roadwidth", "road",
    "접도길이", "전면폭", "필지폭", "frontage", "width", "필지깊이", "대지깊이", "깊이", "depth", "lotdepth",
    "평균높이", "사선평균높이", "높이", "최저높이", "최고높이", "avgheight", "height", "maxheight", "minheight",
    "시도", "광역시도", "sido", "시군구", "군구", "sigungu",
    "읍면동", "법정동", "행정동", "동", "리", "emd",
}


def _pick_header_index(matrix):
    """알려진 헤더 키워드가 가장 많은 행을 헤더로 선택(제목/안내 행·선행 공백 행 무시)."""
    best_i, best_score = None, 0
    for i, row in enumerate(matrix[:25]):
        score = sum(1 for c in row if normalize_header(c) in KNOWN_HEADER_KEYS)
        if score > best_score:
            best_score, best_i = score, i
    if best_i is not None and best_score >= 1:
        return best_i
    # 키워드를 못 찾으면: 첫 비어있지 않은 행을 헤더로
    return next((i for i, row in enumerate(matrix) if any(str(c).strip() for c in row)), None)


def _rows_from_matrix(matrix):
    hi = _pick_header_index(matrix)
    if hi is None:
        return []
    headers = [normalize_header(c) for c in matrix[hi]]
    rows = []
    for raw in matrix[hi + 1:]:
        row = {headers[i]: (str(raw[i]).strip() if i < len(raw) else "")
               for i in range(len(headers)) if headers[i]}
        if any(row.values()):
            rows.append(row)
    return rows


def parse_delimited_table(raw_bytes, suffix):
    text = raw_bytes.decode("utf-8-sig", errors="replace")
    sample = text[:4096]
    delimiter = "\t" if suffix == ".txt" and "\t" in sample else ","
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;")
        delimiter = dialect.delimiter
    except Exception:
        pass
    matrix = [list(r) for r in csv.reader(io.StringIO(text), delimiter=delimiter)]
    return _rows_from_matrix(matrix)


def parse_xlsx_table(raw_bytes):
    ns = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    with zipfile.ZipFile(io.BytesIO(raw_bytes)) as zf:
        shared = []
        if "xl/sharedStrings.xml" in zf.namelist():
            root = ET.fromstring(zf.read("xl/sharedStrings.xml"))
            for si in root.findall("m:si", ns):
                shared.append("".join(t.text or "" for t in si.findall(".//m:t", ns)))
        sheet_name = "xl/worksheets/sheet1.xml"
        if sheet_name not in zf.namelist():
            sheet_name = next((name for name in zf.namelist() if name.startswith("xl/worksheets/sheet")), None)
        if not sheet_name:
            raise ValueError("엑셀 시트가 비어 있습니다.")
        root = ET.fromstring(zf.read(sheet_name))
        matrix = []
        for row in root.findall(".//m:sheetData/m:row", ns):
            values = {}
            max_col = -1
            for cell in row.findall("m:c", ns):
                ref = cell.attrib.get("r", "")
                letters = re.sub(r"[^A-Z]", "", ref.upper())
                col = 0
                for ch in letters:
                    col = col * 26 + ord(ch) - 64
                col -= 1
                max_col = max(max_col, col)
                value_node = cell.find("m:v", ns)
                inline_node = cell.find("m:is/m:t", ns)
                value = ""
                if inline_node is not None:
                    value = inline_node.text or ""
                elif value_node is not None:
                    value = value_node.text or ""
                    if cell.attrib.get("t") == "s":
                        idx = int(float(value))
                        value = shared[idx] if 0 <= idx < len(shared) else ""
                values[col] = str(value).strip()
            if max_col >= 0:
                matrix.append([values.get(i, "") for i in range(max_col + 1)])
    return _rows_from_matrix(matrix)


def parse_block_height_rows(upload):
    name = str(upload.get("name") or "").strip()
    data_url = str(upload.get("data") or "")
    if not name or not data_url:
        raise ValueError("필지 지적 리스트 파일을 첨부해 주세요.")
    if "," in data_url:
        data_url = data_url.split(",", 1)[1]
    raw_bytes = base64.b64decode(data_url)
    suffix = Path(name).suffix.lower()
    if suffix == ".xlsx":
        return parse_xlsx_table(raw_bytes)
    if suffix in {".csv", ".txt"}:
        return parse_delimited_table(raw_bytes, suffix)
    raise ValueError("산정 파일은 .xlsx, .csv, .txt 형식만 지원합니다.")


def calculate_block_height(payload):
    rows = parse_block_height_rows(payload.get("heightListFile") or {})
    default_coverage = (
        parse_float(payload.get("districtPlanCoverage"), None)
        or parse_float(payload.get("ordinanceCoverage"), None)
        or parse_float(payload.get("coverageRatio"), 60.0)
    )
    default_road_width = parse_float(payload.get("roadWidth"), 8.0)
    slope_multiplier = parse_float(payload.get("slopeMultiplier"), 1.5)
    default_frontage = parse_float(payload.get("frontage"), None)
    default_depth = parse_float(payload.get("lotDepth"), None)
    if default_coverage <= 0:
        raise ValueError("건폐율은 0보다 커야 합니다.")
    if slope_multiplier <= 0:
        raise ValueError("도로사선 계수는 0보다 커야 합니다.")

    aliases = {
        "parcel": ["지번", "필지", "필지명", "parcel", "lot", "pnu"],
        "area": ["대지면적", "필지면적", "면적", "area", "lotarea", "sitearea"],
        "coverage": ["건폐율", "coverage", "buildingcoverage", "coverageratio"],
        "road_width": ["도로폭", "도로너비", "접도폭", "도로폭원", "roadwidth", "road"],
        "frontage": ["접도길이", "전면폭", "필지폭", "frontage", "width"],
        "depth": ["필지깊이", "대지깊이", "깊이", "depth", "lotdepth"],
        "avg_height": ["평균높이", "사선평균높이", "높이", "avgheight", "averageheight", "height"],
        "min_height": ["최저높이", "minheight", "lowheight"],
        "max_height": ["최고높이", "maxheight", "highheight"],
    }

    result_rows = []
    total_floor_area = 0.0
    total_volume = 0.0
    for index, row in enumerate(rows, 1):
        parcel = str(row_value(row, aliases["parcel"], f"{index}")).strip()
        area = parse_float(row_value(row, aliases["area"]))
        if not area or area <= 0:
            raise ValueError(f"{index}행의 대지면적/필지면적을 확인해 주세요.")
        coverage = parse_float(row_value(row, aliases["coverage"]), default_coverage)
        if coverage > 1:
            coverage_ratio = coverage / 100.0
        else:
            coverage_ratio = coverage
        if coverage_ratio <= 0:
            raise ValueError(f"{index}행의 건폐율을 확인해 주세요.")
        floor_area = area * coverage_ratio

        explicit_avg = parse_float(row_value(row, aliases["avg_height"]))
        min_height = parse_float(row_value(row, aliases["min_height"]))
        max_height = parse_float(row_value(row, aliases["max_height"]))
        road_width = parse_float(row_value(row, aliases["road_width"]), default_road_width)
        frontage = parse_float(row_value(row, aliases["frontage"]), default_frontage)
        depth = parse_float(row_value(row, aliases["depth"]), default_depth)

        method = "입력 평균높이"
        if explicit_avg is not None:
            avg_height = explicit_avg
            if min_height is None:
                min_height = avg_height
            if max_height is None:
                max_height = avg_height
        elif min_height is not None and max_height is not None:
            avg_height = (min_height + max_height) / 2.0
            method = "최저/최고 평균"
        else:
            if road_width <= 0:
                raise ValueError(f"{index}행은 평균높이 또는 도로폭을 입력해야 합니다.")
            if depth is None:
                if frontage and frontage > 0:
                    depth = floor_area / frontage
                    method = "도로사선 1.5D(접도길이 추정)"
                else:
                    depth = math.sqrt(floor_area)
                    method = "도로사선 1.5D(정방형 추정)"
            else:
                method = "도로사선 1.5D(필지깊이)"
            min_height = slope_multiplier * road_width
            max_height = slope_multiplier * (road_width + depth)
            avg_height = (min_height + max_height) / 2.0

        volume = floor_area * avg_height
        total_floor_area += floor_area
        total_volume += volume
        result_rows.append(
            {
                "parcel": parcel,
                "area": round(area, 3),
                "coverageRatio": round(coverage_ratio * 100, 3),
                "floorArea": round(floor_area, 3),
                "roadWidth": round(road_width, 3) if road_width else None,
                "frontage": round(frontage, 3) if frontage else None,
                "depth": round(depth, 3) if depth else None,
                "minHeight": round(min_height, 3) if min_height is not None else None,
                "maxHeight": round(max_height, 3) if max_height is not None else None,
                "avgHeight": round(avg_height, 3),
                "volume": round(volume, 3),
                "method": method,
            }
        )

    if total_floor_area <= 0:
        raise ValueError("건폐율 적용 바닥면적 합계가 0입니다.")
    block_average = total_volume / total_floor_area
    return {
        "rows": result_rows,
        "totals": {
            "parcelCount": len(result_rows),
            "floorArea": round(total_floor_area, 3),
            "volume": round(total_volume, 3),
            "blockAverageHeight": round(block_average, 3),
        },
    }


def _fmt_num(value, digits=2):
    try:
        return f"{float(value):,.{digits}f}"
    except (TypeError, ValueError):
        return "-"


def _survey_row_address(r):
    """한 행에서 조서용 전체 주소를 조합한다.
    - '주소/소재지지번/address'(지번 포함형)는 그대로 기준으로 사용.
    - '소재지/시도/시군구/읍면동'(지역형) + 별도 '지번' 열이면 합쳐서 완전 주소 생성.
    """
    jibun = str(row_value(r, ["지번", "번지", "parcel", "lot"]) or "").strip()
    full = row_value(r, ["주소", "소재지지번", "소재지번", "address", "addr", "위치"])
    if full:
        addr = str(full).strip()
        # 전체주소 열에 지번이 빠져 있고 별도 지번 열이 있으면 보강
        if jibun and jibun not in addr:
            addr = f"{addr} {jibun}".strip()
        return addr
    region = " ".join(
        str(x).strip() for x in [
            row_value(r, ["시도", "광역시도", "sido"]),
            row_value(r, ["시군구", "군구", "sigungu"]),
            row_value(r, ["소재지", "소재", "읍면동", "법정동", "행정동", "동", "리", "emd"]),
        ] if x and str(x).strip()
    )
    return " ".join(p for p in [region, jibun] if p).strip()


def survey_items_from_file(upload, api_key):
    """지번 목록 엑셀/CSV를 파싱해 조서 대상 필지 목록을 만든다.
    PNU(19자리) 열이 있으면 우선 사용(지번이 이상해도 정확). 없으면 주소→좌표.
    """
    raw_rows = parse_block_height_rows(upload)
    items = []
    for r in raw_rows:
        pnu = re.sub(r"\D", "", str(row_value(r, ["pnu", "고유번호", "필지고유번호", "토지고유번호"]) or ""))
        addr = _survey_row_address(r)
        if len(pnu) == 19:
            # PNU가 가장 확실 → 지오코딩 없이 PNU로 직접 조회
            items.append({"parcel": addr or f"PNU {pnu}", "pnu": pnu})
            continue
        if not addr:
            continue
        item = {"parcel": addr, "pnu": ""}
        try:
            lat, lon, _ = geocode(addr, api_key)
            item["lat"], item["lon"] = lat, lon
        except Exception:
            item["error"] = "주소를 좌표로 변환하지 못했습니다(PNU 열을 넣으면 정확히 찾습니다)"
        items.append(item)
    if not items:
        cols = sorted({k for r in raw_rows for k in r.keys()}) if raw_rows else []
        hint = f" (인식된 열: {', '.join(cols)})" if cols else ""
        raise ValueError(
            "엑셀에서 주소/지번/PNU 열을 찾지 못했습니다. '주소'(또는 '소재지'+'지번') 또는 'PNU' 열이 필요합니다." + hint
        )
    return items


def build_parcel_survey(payload):
    """지도 선택 또는 엑셀 업로드 필지들의 토지 조서(지번·지목·면적·소유구분·용도지역·공시지가)를 생성."""
    api_key = str(payload.get("apiKey") or "").strip()
    if not api_key:
        raise ValueError("필지 조서에는 VWorld API 키가 필요합니다.")
    items = payload.get("parcelListItems") or []
    upload = payload.get("surveyListFile") or {}
    if not items and upload.get("name"):
        items = survey_items_from_file(upload, api_key)
    if not items:
        raise ValueError("지도에서 필지를 선택하거나, 지번 목록 엑셀을 첨부해 주세요.")
    use_char = payload.get("useLandChar", True) is not False
    rows = []
    fail_count = 0
    for idx, item in enumerate(items, 1):
        label = str(item.get("parcel") or "").strip()
        pnu_in = re.sub(r"\D", "", str(item.get("pnu") or ""))
        try:
            feature = None
            if len(pnu_in) == 19:
                # PNU 우선(지번이 이상해도 정확)
                res = get_parcel_feature_by_pnu(pnu_in, api_key)
                feats = (res or {}).get("features") or []
                if feats:
                    feature = feats[0]
                elif not (item.get("lat") is not None and item.get("lon") is not None):
                    raise ValueError(f"PNU {pnu_in}로 필지를 찾지 못했습니다")
            if feature is None:
                if item.get("error"):
                    raise ValueError(item["error"])
                if item.get("lat") is None or item.get("lon") is None:
                    raise ValueError("주소/PNU로 필지를 찾지 못했습니다")
                m_lat = float(item.get("lat"))
                m_lon = float(item.get("lon"))
                feats = (get_parcel_feature(m_lat, m_lon, api_key) or {}).get("features") or []
                if not feats:
                    raise ValueError("해당 위치에서 필지를 찾지 못했습니다")
                feature = feats[0]
            props = feature.get("properties") or {}
            pnu = str(props.get("pnu") or "").strip()
            geom_area = feature_geom_area(feature) or None
            rec = fetch_land_characteristics(pnu, api_key) if (use_char and pnu) else {}
            ladfrl = fetch_land_ladfrl(pnu, api_key) if (use_char and pnu) else {}
            # 소유구분: 세부내용 그대로(개인/법인/국유지/도유지/군유지/외국인/종중 등)
            owner = owner_raw = str(ladfrl.get("posesnSeCodeNm") or "").strip()
            cnrs = parse_float(ladfrl.get("cnrsPsnCo"))
            if owner and cnrs and cnrs > 1:
                owner = f"{owner}(공유 {int(cnrs)}인)"
            reg_area = parse_float(rec.get("lndpclAr"))
            jiga = parse_float(rec.get("pblntfPclnd"))
            z1 = str(rec.get("prposArea1Nm") or "").strip()
            z2 = str(rec.get("prposArea2Nm") or "").strip()
            zone = z1 + (f" / {z2}" if z2 and z2 != "지정되지않음" else "")
            jibun_raw = str(props.get("jibun") or "").strip()      # 예: "161-7 대"
            parts = jibun_raw.split()
            jibun_no = parts[0] if parts else f"{props.get('bonbun', '')}-{props.get('bubun', '')}".strip("-")
            jimok = str(rec.get("lndcgrCodeNm") or (parts[1] if len(parts) > 1 else "")).strip()
            sido = str(props.get("ctp_nm") or "").strip()
            sgg = str(props.get("sig_nm") or "").strip()
            emd = str(props.get("emd_nm") or "").strip()
            addr = str(props.get("addr") or "").strip() or " ".join(x for x in [sido, sgg, emd, jibun_no] if x)
            area_for_value = reg_area if reg_area else geom_area
            area_diff_pct = round((geom_area - reg_area) / reg_area * 100, 1) if (reg_area and geom_area) else None
            rows.append({
                "no": idx,
                "areaDiffPct": area_diff_pct,
                "addr": addr, "sido": sido, "sigungu": sgg, "emd": emd,
                "jibun": jibun_no, "jimok": jimok or "-", "pnu": pnu,
                "owner": owner or "-", "ownerRaw": owner_raw,
                "regArea": round(reg_area, 2) if reg_area else None,
                "geomArea": round(geom_area, 2) if geom_area else None,
                "zone": zone or "-",
                "jiga": int(round(jiga)) if jiga else None,
                "jigaTotal": int(round(jiga * area_for_value)) if (jiga and area_for_value) else None,
                "useSittn": str(rec.get("ladUseSittnNm") or "").strip(),
                "year": str(rec.get("stdrYear") or "").strip(),
                "error": "",
            })
        except Exception as exc:
            fail_count += 1
            rows.append({
                "no": idx, "areaDiffPct": None,
                "addr": label or "(주소 미상)", "sido": "", "sigungu": "", "emd": "",
                "jibun": "", "jimok": "-", "pnu": "",
                "owner": "-", "ownerRaw": "",
                "regArea": None, "geomArea": None, "zone": "-",
                "jiga": None, "jigaTotal": None, "useSittn": "", "year": "",
                "error": str(exc)[:80] or "조회 실패",
            })
    totals = {
        "count": len(rows),
        "failCount": fail_count,
        "source": "file" if (not payload.get("parcelListItems") and upload.get("name")) else "map",
        "regArea": round(sum(r["regArea"] or 0 for r in rows), 2),
        "geomArea": round(sum(r["geomArea"] or 0 for r in rows), 2),
        "jigaTotal": sum(r["jigaTotal"] or 0 for r in rows),
    }
    return {"rows": rows, "totals": totals}


def build_parcel_survey_xlsx(payload):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    rows = payload.get("rows") or []
    totals = payload.get("totals") or {}
    if not rows:
        raise ValueError("조서로 만들 선택 필지가 없습니다. 먼저 필지를 선택해 주세요.")
    wb = Workbook()
    ws = wb.active
    ws.title = "필지조서"
    title_font = Font(bold=True, size=14, color="14301F")
    legend_font = Font(size=9, color="555555")
    head_font = Font(bold=True, color="FFFFFF", size=10)
    head_fill = PatternFill("solid", fgColor="1B4F9C")
    total_fill = PatternFill("solid", fgColor="EAF3EC")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center")
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    num_fmt = "#,##0.00"
    int_fmt = "#,##0"

    ws["A1"] = "필지 조서(토지조서)"
    ws["A1"].font = title_font
    ws["A2"] = f"생성: {time.strftime('%Y-%m-%d %H:%M')}   ·   필지 수 = {len(rows)}   ·   출처: VWorld 연속지적·토지특성"
    ws["A2"].font = legend_font
    ws["A3"] = "※ 공부면적·지목·용도지역·개별공시지가는 토지특성(기준연도) 기준, 도형면적은 연속지적 폴리곤 계산값. 공시지가액 = (공부면적 있으면 공부, 없으면 도형) × 개별공시지가."
    ws["A3"].font = legend_font

    header_row = 5
    headers = ["연번", "소재지", "지번", "지목", "소유구분", "공부면적(㎡)", "도형면적(㎡)",
               "도형-공부 차(%)", "용도지역", "개별공시지가(원/㎡)", "공시지가액(원)",
               "이용상황", "기준연도", "PNU", "비고"]
    for c, text in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=c, value=text)
        cell.font = head_font
        cell.fill = head_fill
        cell.alignment = center
        cell.border = border
    num_cols = {6, 7, 10, 11}   # 공부면적, 도형면적, 개별공시지가, 공시지가액
    diff_font = Font(color="C0392B", bold=True)
    for i, row in enumerate(rows):
        rr = header_row + 1 + i
        diff = parse_float(row.get("areaDiffPct"))
        big_diff = diff is not None and abs(diff) >= 10
        vals = [row.get("no"), row.get("addr"), row.get("jibun"), row.get("jimok"), row.get("owner"),
                parse_float(row.get("regArea")), parse_float(row.get("geomArea")), diff,
                row.get("zone"), row.get("jiga"), row.get("jigaTotal"),
                row.get("useSittn"), row.get("year"), row.get("pnu"), row.get("error") or ""]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=rr, column=c, value=v)
            cell.border = border
            if c in num_cols:
                cell.number_format = num_fmt if c in (6, 7) else int_fmt
                cell.alignment = right
            elif c == 8:                       # 도형-공부 차(%)
                cell.number_format = '+0.0;-0.0'
                cell.alignment = right
                if big_diff:
                    cell.font = diff_font
            elif c in (1, 5):
                cell.alignment = center
            else:
                cell.alignment = left
            if big_diff and c in (6, 7):
                cell.font = diff_font
    tr = header_row + 1 + len(rows)
    ws.cell(row=tr, column=1, value="합계").font = Font(bold=True)
    ws.cell(row=tr, column=6, value=parse_float(totals.get("regArea")))
    ws.cell(row=tr, column=7, value=parse_float(totals.get("geomArea")))
    ws.cell(row=tr, column=11, value=totals.get("jigaTotal"))
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=tr, column=c)
        cell.border = border
        cell.fill = total_fill
        if c in (6, 7):
            cell.number_format = num_fmt
            cell.font = Font(bold=True)
            cell.alignment = right
        if c == 11:
            cell.number_format = int_fmt
            cell.font = Font(bold=True)
            cell.alignment = right
    widths = [6, 30, 10, 7, 11, 12, 12, 13, 20, 16, 16, 12, 9, 22, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_height_report_xlsx(payload):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    rows = payload.get("rows") or []
    if not rows:
        raise ValueError("보고서로 만들 산정 결과가 없습니다. 먼저 평균높이를 산정해 주세요.")
    slope = parse_float(payload.get("slopeMultiplier"), 1.5)
    totals = payload.get("totals") or {}

    wb = Workbook()
    ws = wb.active
    ws.title = "평균높이산정"

    bold = Font(bold=True)
    title_font = Font(bold=True, size=14, color="14301F")
    legend_font = Font(size=9, color="555555")
    head_font = Font(bold=True, color="FFFFFF", size=10)
    head_fill = PatternFill("solid", fgColor="1B4F9C")
    input_fill = PatternFill("solid", fgColor="FFF4E0")   # 입력/원천값
    calc_fill = PatternFill("solid", fgColor="E7EFFB")    # 자동계산(수식)
    total_fill = PatternFill("solid", fgColor="EAF3EC")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right = Alignment(horizontal="right", vertical="center")
    num_fmt = "#,##0.00"

    ws["A1"] = "블록별 평균높이 산정 보고서"
    ws["A1"].font = title_font
    ws["A2"] = f"생성: {time.strftime('%Y-%m-%d %H:%M')}   ·   사선계수(도로사선) = {slope}   ·   필지 수 = {len(rows)}"
    ws["A2"].font = legend_font
    legends = [
        "[산식] 바닥면적 = 대지면적 × 건폐율(%) ÷ 100   (대지면적은 공부면적(토지특성) 있으면 공부, 없으면 도형면적 사용)",
        "[산식] (1.5D 도로사선) 최저높이 = 사선계수 × (도로폭 + 전면건축선후퇴)",
        "[산식] 최고높이 = 사선계수 × (도로폭 + 전면건축선후퇴 + (필지깊이 − 후면건축선후퇴))",
        "[산식] 평균높이 = (최저높이 + 최고높이) ÷ 2,   체적 = 바닥면적 × 평균높이,   블록 평균높이 = Σ체적 ÷ Σ바닥면적",
        "[입지] 도로폭 20m 이상=간선부, 미만=이면부.  [최고높이상한] 용도지역×입지 기준값을 중심지 위계(±%)로 조정한 값.",
        "※ 주황색 셀 = 입력/원천값,  파란색 셀 = Excel 수식(셀 클릭 시 산식 표시, 입력값을 바꾸면 자동 재계산).  '검산' 열은 숫자를 대입한 식입니다.",
    ]
    r = 4
    for text in legends:
        ws.cell(row=r, column=1, value=text).font = legend_font
        r += 1

    header_row = r + 1
    headers = [
        "필지", "대지면적(㎡)", "건폐율(%)", "바닥면적(㎡)", "도로폭(m)", "접도길이(m)",
        "필지깊이(m)", "사선계수", "최저높이(m)", "최고높이(m)", "평균높이(m)", "체적(㎥)",
        "산정방식", "건폐율출처", "신뢰도", "검산(산식·숫자 대입)", "비고",
        "전면후퇴(m)", "후면후퇴(m)", "입지", "용도지역상한(m)", "위계조정후상한(m)", "상한판정",
        "공부면적(㎡)", "도형면적(㎡)", "면적출처",
    ]
    ncols = len(headers)
    for c, text in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=c, value=text)
        cell.font = head_font
        cell.fill = head_fill
        cell.alignment = center
        cell.border = border

    first_data = header_row + 1
    front_col = get_column_letter(18)
    rear_col = get_column_letter(19)
    input_cols = {2, 3, 5, 6, 7, 8, 9, 10, 18, 19}
    calc_cols = {4, 11, 12}
    num_cols = {2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 18, 19, 21, 22, 24, 25}

    for i, row in enumerate(rows):
        rr = first_data + i
        area = parse_float(row.get("area"))
        cov = parse_float(row.get("coverageRatio"))
        rw = parse_float(row.get("roadWidth"))
        fr = parse_float(row.get("frontage"))
        dp = parse_float(row.get("depth"))
        mn = parse_float(row.get("minHeight"))
        mx = parse_float(row.get("maxHeight"))
        av = parse_float(row.get("avgHeight"))
        fl = parse_float(row.get("floorArea"))
        vol = parse_float(row.get("volume"))
        method = str(row.get("method") or "")
        fsb = parse_float(row.get("frontSetback")) or 0.0
        rsb = parse_float(row.get("rearSetback")) or 0.0
        roadsine = bool(
            rw and dp is not None and mn is not None and mx is not None
            and abs(mn - slope * (rw + fsb)) < 0.06
            and abs(mx - slope * (rw + fsb + max(dp - rsb, 0.0))) < 0.06
        )

        ws.cell(row=rr, column=1, value=row.get("parcel") or f"{i + 1}")
        ws.cell(row=rr, column=2, value=area)
        ws.cell(row=rr, column=3, value=cov)
        ws.cell(row=rr, column=4, value=f"=B{rr}*C{rr}/100")
        ws.cell(row=rr, column=5, value=rw)
        ws.cell(row=rr, column=6, value=fr)
        ws.cell(row=rr, column=7, value=dp)
        ws.cell(row=rr, column=8, value=(slope if roadsine else None))
        if roadsine:
            # 최저=계수×(도로폭+전면후퇴),  최고=계수×(도로폭+전면후퇴+(깊이−후면후퇴))
            ws.cell(row=rr, column=9, value=f"=H{rr}*(E{rr}+{front_col}{rr})")
            ws.cell(row=rr, column=10, value=f"=H{rr}*(E{rr}+{front_col}{rr}+MAX(G{rr}-{rear_col}{rr},0))")
        else:
            ws.cell(row=rr, column=9, value=mn)
            ws.cell(row=rr, column=10, value=mx)
        ws.cell(row=rr, column=11, value=f"=(I{rr}+J{rr})/2")
        ws.cell(row=rr, column=12, value=f"=D{rr}*K{rr}")
        ws.cell(row=rr, column=13, value=method)
        ws.cell(row=rr, column=14, value=row.get("coverageSource") or "")
        ws.cell(row=rr, column=15, value=row.get("confidence") or "")

        checks = []
        if row.get("zoneUpgraded"):
            checks.append(f"용도지역 상향 {row.get('zoneAuto') or '-'}→{row.get('zone') or '-'}(건폐율·최고높이 상한 상향 적용)")
        if fsb > 0 or rsb > 0:
            checks.append(f"건축선후퇴 전면{_fmt_num(fsb)}/후면{_fmt_num(rsb)}m(기준거리=도로폭+전면후퇴, 깊이는 후면후퇴 차감)")
        if area is not None and cov is not None:
            checks.append(f"바닥면적={_fmt_num(area)}×{_fmt_num(cov)}%={_fmt_num(fl)}")
        if roadsine:
            checks.append(f"최저={_fmt_num(slope)}×({_fmt_num(rw)}+{_fmt_num(fsb)})={_fmt_num(mn)}")
            checks.append(f"최고={_fmt_num(slope)}×({_fmt_num(rw)}+{_fmt_num(fsb)}+({_fmt_num(dp)}−{_fmt_num(rsb)}))={_fmt_num(mx)}")
        if mn is not None and mx is not None:
            checks.append(f"평균=({_fmt_num(mn)}+{_fmt_num(mx)})/2={_fmt_num(av)}")
        if fl is not None and av is not None:
            checks.append(f"체적={_fmt_num(fl)}×{_fmt_num(av)}={_fmt_num(vol)}")
        ws.cell(row=rr, column=16, value=" · ".join(checks))
        ws.cell(row=rr, column=17, value=row.get("note") or "")

        # 입지·최고높이 상한(②·①)
        cap_base = parse_float(row.get("heightCapBase"))
        cap_adj = parse_float(row.get("heightCapAdjusted"))
        ws.cell(row=rr, column=18, value=fsb)
        ws.cell(row=rr, column=19, value=rsb)
        ws.cell(row=rr, column=20, value=row.get("arterialClass") or "-")
        ws.cell(row=rr, column=21, value=cap_base)
        ws.cell(row=rr, column=22, value=cap_adj)
        if cap_adj is not None:
            ws.cell(row=rr, column=23, value=("초과" if row.get("capExceeded") else "적합"))
        else:
            ws.cell(row=rr, column=23, value="대상外")
        # 공부면적/도형면적/면적출처(대지면적 B열은 산정에 쓴 값 = 공부 우선)
        ws.cell(row=rr, column=24, value=parse_float(row.get("regArea")))
        ws.cell(row=rr, column=25, value=parse_float(row.get("geomArea")))
        ws.cell(row=rr, column=26, value=row.get("areaSource") or "")

        for c in range(1, ncols + 1):
            cell = ws.cell(row=rr, column=c)
            cell.border = border
            if c in num_cols:
                cell.number_format = num_fmt
                cell.alignment = right
            else:
                cell.alignment = left
            if c in input_cols and not (c in (9, 10) and roadsine):
                cell.fill = input_fill
            if c in calc_cols or (c in (9, 10) and roadsine):
                cell.fill = calc_fill

    last_data = first_data + len(rows) - 1
    tr = last_data + 1
    ws.cell(row=tr, column=1, value="합계 / 블록 평균").font = bold
    ws.cell(row=tr, column=4, value=f"=SUM(D{first_data}:D{last_data})")
    ws.cell(row=tr, column=11, value=f"=L{tr}/D{tr}")
    ws.cell(row=tr, column=12, value=f"=SUM(L{first_data}:L{last_data})")
    ws.cell(
        row=tr, column=16,
        value=(
            f"블록평균높이=Σ체적÷Σ바닥면적={_fmt_num(totals.get('volume'))}"
            f"÷{_fmt_num(totals.get('floorArea'))}={_fmt_num(totals.get('blockAverageHeight'))}"
        ),
    )
    for c in range(1, ncols + 1):
        cell = ws.cell(row=tr, column=c)
        cell.border = border
        cell.fill = total_fill
        if c in (4, 11, 12):
            cell.number_format = num_fmt
            cell.font = bold
            cell.alignment = right

    widths = [16, 12, 9, 12, 9, 11, 11, 9, 11, 11, 11, 12, 20, 12, 8, 46, 18,
              11, 11, 8, 14, 16, 9, 12, 12, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = ws.cell(row=first_data, column=1)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def start_cadastre_job(payload):
    address = payload.get("address", "").strip()
    api_key = payload.get("apiKey", "").strip()
    radius = str(payload.get("radius") or "1000").strip()
    vworld_id = payload.get("vworldId", "").strip()
    vworld_pw = payload.get("vworldPassword", "")
    out_dir = Path(payload.get("outDir") or DEFAULT_OUT_DIR)
    out_dir.mkdir(parents=True, exist_ok=True)
    parcel_list_paths = save_parcel_list_upload(payload, out_dir)
    log_dir = out_dir / "logs"
    log_dir.mkdir(parents=True, exist_ok=True)
    log_path = log_dir / f"html_run_{time.strftime('%Y%m%d_%H%M%S')}.log"

    if not CADASTRE_SCRIPT.exists():
        raise FileNotFoundError(f"연속지적 스크립트를 찾지 못했습니다: {CADASTRE_SCRIPT}")

    args = [
        "powershell.exe",
        "-ExecutionPolicy",
        "Bypass",
        "-File",
        str(CADASTRE_SCRIPT),
        "-Address",
        address,
        "-VWorldKey",
        api_key,
        "-Radius",
        radius,
        "-OutDir",
        str(out_dir),
    ]
    for parcel_list_path in parcel_list_paths:
        args += ["-ParcelListPath", str(parcel_list_path)]
    if vworld_id and vworld_pw:
        args += ["-VWorldId", vworld_id, "-VWorldPassword", vworld_pw]

    log_file = log_path.open("w", encoding="utf-8")
    process = subprocess.Popen(args, cwd=str(APP_DIR), stdout=log_file, stderr=subprocess.STDOUT, text=True)
    job_id = str(int(time.time() * 1000))
    JOBS[job_id] = {"process": process, "log": log_path, "outDir": out_dir}
    return job_id, log_path, out_dir


class BridgeHandler(BaseHTTPRequestHandler):
    def log_message(self, format, *args):
        return

    def do_OPTIONS(self):
        cors_headers(self)
        self.end_headers()

    def _set_request_domain(self):
        # 페이지 접속 도메인을 VWorld WFS DOMAIN으로 사용(포트 제외).
        host = self.headers.get("Host") or ""
        origin = self.headers.get("Origin") or self.headers.get("Referer") or ""
        if origin:
            try:
                host = urllib.parse.urlparse(origin).netloc or host
            except Exception:
                pass
        _request_ctx.domain = host.split(":")[0].strip()

    def do_GET(self):
        self._set_request_domain()
        parsed = urllib.parse.urlparse(self.path)
        path = parsed.path
        qs = urllib.parse.parse_qs(parsed.query)
        try:
            if path == "/":
                self.serve_file(DEFAULT_HTML)
                return
            if path == "/api/health":
                json_response(self, {"ok": True})
                return
            if path == "/api/geocode":
                address = qs.get("address", [""])[0].strip()
                api_key = qs.get("apiKey", [""])[0].strip()
                lat, lon, refined = geocode(address, api_key)
                jibun_addr, road_addr = reverse_geocode_both(lat, lon, api_key)
                json_response(self, {"ok": True, "lat": lat, "lon": lon, "refined": refined,
                                     "jibunAddr": jibun_addr, "roadAddr": road_addr})
                return
            if path == "/api/zoning-legend":
                api_key = qs.get("apiKey", [""])[0].strip()
                try:
                    bbox = [float(x) for x in qs.get("bbox", [""])[0].split(",")]
                except Exception:
                    bbox = []
                if len(bbox) != 4:
                    json_response(self, {"ok": False, "error": "bbox(minLon,minLat,maxLon,maxLat)가 필요합니다."}, status=400)
                    return
                items = sample_zoning_legend(api_key, bbox)
                json_response(self, {"ok": True, "items": items})
                return
            if path == "/api/reverse-parcel":
                lat = float(qs.get("lat", ["0"])[0])
                lon = float(qs.get("lon", ["0"])[0])
                api_key = qs.get("apiKey", [""])[0].strip()
                parcel = reverse_geocode_parcel(lat, lon, api_key)
                json_response(self, {"ok": True, "parcel": parcel})
                return
            if path == "/api/parcel-feature":
                lat = float(qs.get("lat", ["0"])[0])
                lon = float(qs.get("lon", ["0"])[0])
                api_key = qs.get("apiKey", [""])[0].strip()
                feature = get_parcel_feature(lat, lon, api_key)
                json_response(self, {"ok": True, "feature": feature})
                return
            if path == "/api/context-facilities":
                lat = float(qs.get("lat", ["0"])[0])
                lon = float(qs.get("lon", ["0"])[0])
                radius = float(qs.get("radius", ["2500"])[0])
                facilities = get_context_facilities(lat, lon, radius)
                json_response(self, {"ok": True, "facilities": facilities})
                return
            if path == "/api/cadastre-dxf":
                address = qs.get("address", [""])[0].strip()
                api_key = qs.get("apiKey", [""])[0].strip()
                radius = qs.get("radius", ["300"])[0]
                domain = qs.get("domain", [""])[0].strip()
                result = build_cadastre_dxf(address, api_key, radius, domain=domain)
                data = result["data"]
                cors_headers(self, content_type="application/dxf")
                fname = urllib.parse.quote(f"연속지적_{result['refined']}_{result['radius']}m.dxf".replace(" ", "_"))
                self.send_header("Content-Disposition", f"attachment; filename*=UTF-8''{fname}")
                self.send_header("X-Parcel-Count", str(result["parcelCount"]))
                self.send_header("X-Zone-Count", str(result["zoneCount"]))
                self.send_header("X-Plan-Count", str(result.get("planCount", 0)))
                self.send_header("X-Truncated", "1" if result["truncated"] else "0")
                self.send_header("X-Errors", urllib.parse.quote(" | ".join(result.get("errors") or [])[:400]))
                self.send_header("Content-Length", str(len(data)))
                self.end_headers()
                self.wfile.write(data)
                return
            if path == "/api/satellite-map":
                address = qs.get("address", [""])[0].strip()
                api_key = qs.get("apiKey", [""])[0].strip()
                radius = float(qs.get("radius", ["3000"])[0])
                show_radius = qs.get("showRadius", ["1"])[0] != "0"
                show_facilities = qs.get("facilities", ["1"])[0] != "0"
                map_type = qs.get("mapType", ["satellite"])[0]
                if map_type not in {"satellite", "street", "blank"}:
                    map_type = "satellite"
                size = qs.get("size", ["high"])[0]
                dimensions = {
                    "standard": (1600, 1000),
                    "high": (3200, 2000),
                    "ultra": (4800, 3000),
                }.get(size, (3200, 2000))
                category_text = qs.get("categories", [""])[0]
                categories = [item for item in category_text.split(",") if item] if category_text else None
                out, lat, lon, refined = make_satellite_map(address, api_key, width=dimensions[0], height=dimensions[1], radius=radius, show_facilities=show_facilities, categories=categories, map_type=map_type, show_radius=show_radius)
                data = out.read_bytes()
                cors_headers(self, content_type="image/png")
                filename = urllib.parse.quote(out.name)
                self.send_header("Content-Disposition", f"attachment; filename*=UTF-8''{filename}")
                self.send_header("Content-Length", str(len(data)))
                self.end_headers()
                self.wfile.write(data)
                return
            if path == "/api/status":
                job_id = qs.get("jobId", [""])[0]
                job = JOBS.get(job_id)
                if not job:
                    json_response(self, {"ok": False, "error": "작업을 찾지 못했습니다."}, status=404)
                    return
                process = job["process"]
                log_path = job["log"]
                log = log_path.read_text(encoding="utf-8", errors="replace") if log_path.exists() else ""
                json_response(
                    self,
                    {
                        "ok": True,
                        "running": process.poll() is None,
                        "exitCode": process.poll(),
                        "log": log[-6000:],
                        "logPath": str(log_path),
                        "outDir": str(job["outDir"]),
                    },
                )
                return
            self.serve_file(ROOT / path.lstrip("/"))
        except Exception as exc:
            json_response(self, {"ok": False, "error": str(exc)}, status=500)

    def do_POST(self):
        self._set_request_domain()
        parsed = urllib.parse.urlparse(self.path)
        try:
            length = int(self.headers.get("Content-Length", "0"))
            payload = json.loads(self.rfile.read(length).decode("utf-8")) if length else {}
            if parsed.path == "/api/run-cadastre":
                if WEB_MODE:
                    json_response(self, {"ok": False, "error": "연속지적 실행(QGIS·다운로드)은 데스크톱 실행 전용입니다. 바탕화면 아이콘으로 실행해 주세요."}, status=400)
                    return
                job_id, log_path, out_dir = start_cadastre_job(payload)
                json_response(self, {"ok": True, "jobId": job_id, "logPath": str(log_path), "outDir": str(out_dir)})
                return
            if parsed.path == "/api/calculate-block-height":
                result = calculate_block_height(payload)
                json_response(self, {"ok": True, **result})
                return
            if parsed.path == "/api/height-report":
                data = build_height_report_xlsx(payload)
                cors_headers(self, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                filename = urllib.parse.quote(f"block_height_report_{time.strftime('%Y%m%d_%H%M%S')}.xlsx")
                self.send_header("Content-Disposition", f"attachment; filename*=UTF-8''{filename}")
                self.send_header("Content-Length", str(len(data)))
                self.end_headers()
                self.wfile.write(data)
                return
            if parsed.path == "/api/analyze-selected-parcels":
                result = analyze_selected_parcels(payload)
                json_response(self, {"ok": True, **result})
                return
            if parsed.path == "/api/parcel-survey":
                result = build_parcel_survey(payload)
                json_response(self, {"ok": True, **result})
                return
            if parsed.path == "/api/parcel-survey-report":
                data = build_parcel_survey_xlsx(payload)
                cors_headers(self, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                filename = urllib.parse.quote(f"parcel_survey_{time.strftime('%Y%m%d_%H%M%S')}.xlsx")
                self.send_header("Content-Disposition", f"attachment; filename*=UTF-8''{filename}")
                self.send_header("Content-Length", str(len(data)))
                self.end_headers()
                self.wfile.write(data)
                return
            if parsed.path == "/api/select-folder":
                if WEB_MODE:
                    json_response(self, {"ok": False, "error": "폴더 선택은 데스크톱 실행 전용입니다."}, status=400)
                    return
                path = select_folder(payload.get("initialDir", ""))
                json_response(self, {"ok": True, "path": path})
                return
            json_response(self, {"ok": False, "error": "알 수 없는 요청입니다."}, status=404)
        except Exception as exc:
            json_response(self, {"ok": False, "error": str(exc)}, status=500)

    def serve_file(self, path):
        try:
            resolved = Path(path).resolve()
        except (OSError, ValueError):
            json_response(self, {"ok": False, "error": "파일을 찾지 못했습니다."}, status=404)
            return
        if not resolved.is_relative_to(ROOT):
            json_response(self, {"ok": False, "error": "파일을 찾지 못했습니다."}, status=404)
            return
        path = resolved
        if not path.exists() or not path.is_file():
            json_response(self, {"ok": False, "error": "파일을 찾지 못했습니다."}, status=404)
            return
        content_type = mimetypes.guess_type(path.name)[0] or "application/octet-stream"
        if content_type.startswith("text/"):
            content_type = f"{content_type}; charset=utf-8"
        data = path.read_bytes()
        self.send_response(200)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def log_message(self, *_):
        return


def main():
    # 배포 환경: PORT 환경변수 사용 + 0.0.0.0 바인딩. 로컬: 127.0.0.1:8788.
    env_port = os.environ.get("PORT")
    port = int(env_port) if env_port else (int(sys.argv[1]) if len(sys.argv) > 1 else 8788)
    host = "0.0.0.0" if (env_port or os.name != "nt") else "127.0.0.1"
    server = ThreadingHTTPServer((host, port), BridgeHandler)
    if sys.stdout:
        print(f"Cadastre bridge running: http://{host}:{port}/", flush=True)
    server.serve_forever()


if __name__ == "__main__":
    main()
